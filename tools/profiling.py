import openpyxl
from langchain_core.tools import tool


@tool
def profile_file(path: str, sheet_name: str = "") -> dict:
    """Read a file, detect its structure, profile columns, and find category hierarchy.
    
    This is the first tool to call. It returns headers, column profiles, sample rows,
    row count, category candidates, and file fingerprint in one shot."""
    import json
    from helpers import read_file, get_headers_and_data, fingerprint_headers, load_cached_mapping

    if not sheet_name:
        result = detect_data_sheet.invoke({"path": path})
        sheet_name = result["sheet"]

    rows = read_file(path, sheet_name)
    hr = 0
    dr = 1

    first_15 = [{f"col_{j}": str(c)[:40] for j, c in enumerate(row[:20]) if c is not None and str(c).strip()} for row in rows[:15]]
    header_prompt = f"""Given the first 15 rows of a spreadsheet, identify:
1. Which row index (0-based) contains the column headers
2. Which row index does the actual product data start at

Rows:
{json.dumps(first_15, indent=2)}
Return JSON: {{"header_row": int, "data_start_row": int}}"""

    from google import genai
    import os
    hclient = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
    hresp = hclient.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=header_prompt,
        config={"response_mime_type": "application/json"}
    )
    hresult = json.loads(hresp.text)
    if isinstance(hresult, list):
        hresult = hresult[0] if hresult else {}
    hr = hresult.get("header_row", 0)
    dr = hresult.get("data_start_row", hr + 1)
    if dr < hr + 1:
        dr = hr + 1

    headers, data = get_headers_and_data(rows, hr)
    data = rows[dr:]
    fp = fingerprint_headers(headers)
    cached = load_cached_mapping(fp)

    cols = []
    for ci, h in enumerate(headers):
        vals = []
        for row in data:
            if ci < len(row) and row[ci] is not None and str(row[ci]).strip():
                vals.append(str(row[ci]))
        uniq = list(set(vals))
        sample = (uniq[:3] + uniq[-2:]) if len(uniq) > 5 else uniq
        cols.append({
            "name": h,
            "non_null": len(vals),
            "unique": len(uniq),
            "sample": sample,
            "unique_values": uniq if len(uniq) <= 100 else []
        })

    cats = detect_category_structure.invoke({"path": path, "data_sheet": sheet_name})

    meta = [{headers[j]: str(rows[mr][j])[:60] for j in range(min(len(headers), len(rows[mr]))) if rows[mr][j] is not None and str(rows[mr][j]).strip()} for mr in range(hr)]

    return {
        "fingerprint": fp,
        "sheet_name": sheet_name,
        "header_row": hr,
        "data_start_row": dr,
        "headers": headers,
        "profiles": cols,
        "row_count": len(data),
        "sample_rows": [{headers[j]: str(r[j])[:60] for j in range(min(len(headers), len(r)))} for r in data[:3]],
        "metadata": meta,
        "category_candidates": cats,
        "cached_mapping_exists": cached is not None
    }


@tool
def detect_data_sheet(path: str) -> dict:
    """Scan all sheets in a workbook and return the one with the most data cells.
    
    Returns the sheet name and cell count of the best candidate.
    Call this when no sheet name is provided."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    best = wb.sheetnames[0]
    best_size = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        first = next(ws.iter_rows(max_row=1, values_only=True), [])
        cols = sum(1 for c in first if c is not None)
        rows = sum(1 for _ in ws.iter_rows())
        size = cols * rows
        if size > best_size:
            best_size = size
            best = sn
    wb.close()
    return {"sheet": best, "cells": best_size}


@tool
def profile_columns(headers: list[str], rows: list[list]) -> list[dict]:
    """Scan columns and return stats: name, non_null count, unique count, sample values.
    
    Use this to understand the shape and content of source data before mapping."""
    profiles = []
    for ci, h in enumerate(headers):
        vals = []
        for row in rows:
            if ci < len(row) and row[ci] is not None and str(row[ci]).strip():
                vals.append(str(row[ci]))
        uniq = list(set(vals))
        if len(uniq) > 5:
            sample = uniq[:3] + uniq[-2:]
        else:
            sample = uniq
        profiles.append({
            "name": h,
            "non_null": len(vals),
            "unique": len(uniq),
            "sample": sample,
            "unique_values": uniq if len(uniq) <= 100 else []
        })
    return profiles


@tool
def detect_category_structure(path: str, data_sheet: str) -> list[dict]:
    """Find sheets that may contain category hierarchy data.

    Returns a list of candidates. Each candidate has sheet name, column headers,
    and the first 3 data rows. The caller (LLM) decides which columns form the path."""
    candidates = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if sn == data_sheet:
                continue
            ws = wb[sn]
            first = next(ws.iter_rows(max_row=1, values_only=True), [])
            headers = [str(c) if c is not None else "" for c in first]
            non_empty = sum(1 for h in headers if h.strip())
            if non_empty < 2:
                continue
            sample = []
            for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
                sample.append({headers[i]: str(row[i]) if i < len(row) and row[i] is not None else "" for i in range(len(headers))})
            candidates.append({
                "sheet": sn,
                "headers": headers,
                "rows": sample
            })
        wb.close()
    except Exception:
        pass
    return candidates
