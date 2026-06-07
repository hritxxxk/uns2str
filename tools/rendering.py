import os

from langchain_core.tools import tool
import openpyxl
from openpyxl import Workbook


@tool
def render_category_xlsx(paths: list[str], blank_path: str = "") -> Workbook:
    """Generate category master workbook. Writes into blank template if provided."""
    if blank_path and os.path.exists(blank_path):
        wb = openpyxl.load_workbook(blank_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        for i, p in enumerate(paths, 2):
            ws.cell(row=i, column=1, value=p)
        return wb
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Category Path"
    for i, p in enumerate(paths, 2):
        ws.cell(row=i, column=1, value=p)
    return wb


@tool
def render_attribute_xlsx(defs: list[dict], blank_path: str = "") -> Workbook:
    """Generate attribute master workbook. Writes into PIM blank template if provided."""
    OURS = [
        "Attribute Name", "Short Name", "Display Name", "Attribute Type",
        "Attribute Data Type", "Constraint", "Length", "Mandatory",
        "Filter", "Editability", "Visibility", "Searchable", "Auto Translate",
        "Attribute Group", "Reference Master", "Reference Attribute", "Status"
    ]

    if blank_path and os.path.exists(blank_path):
        wb = openpyxl.load_workbook(blank_path)
        ws = wb.active
        pim_headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        for ci, ph in enumerate(pim_headers, 1):
            normalized = ph.replace("*", "").strip().lower().replace(" ", "_")
            if normalized in [c.lower().replace(" ", "_") for c in OURS]:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.value = None
                break
        col_map = {}
        for ci, ph in enumerate(pim_headers, 1):
            key = ph.replace("*", "").strip().lower().replace(" ", "_")
            ours_key = next((k for k in [c.lower().replace(" ", "_") for c in OURS] if k == key), None)
            if ours_key:
                col_map[ci] = ours_key
        for ri, d in enumerate(defs, 2):
            for ci, key in col_map.items():
                val = d.get(key)
                if val is not None:
                    ws.cell(row=ri, column=ci, value=val)
        return wb

    wb = Workbook()
    ws = wb.active
    for i, c in enumerate(OURS, 1):
        ws.cell(row=1, column=i, value=c)
    for ri, d in enumerate(defs, 2):
        for ci, c in enumerate(OURS, 1):
            key = c.lower().replace(" ", "_")
            val = d.get(key)
            if val is not None:
                ws.cell(row=ri, column=ci, value=val)
    return wb


@tool
def render_reference_xlsx(refs: dict[str, list[str]]) -> Workbook:
    """Generate reference master workbook. One column per attribute master."""
    wb = Workbook()
    ws = wb.active
    all_values = list(refs.values())
    max_rows = max(len(v) for v in all_values) if all_values else 0
    for ci, (master, values) in enumerate(refs.items(), 1):
        ws.cell(row=1, column=ci, value=master)
        for ri, v in enumerate(values, 2):
            ws.cell(row=ri, column=ci, value=v)
    return wb


@tool
def render_all_templates(fingerprint: str, category_hierarchy: list, attribute_definitions: list,
                         reference_values: dict, headers: list, product_rows: list,
                         attr_names: list, blank_category_path: str = "",
                         blank_attribute_path: str = "") -> dict:
    """Generate all 4 PIM output templates and save them to the output directory.
    
    Returns a dict of {template_type: file_path}."""
    import os
    os.makedirs("output", exist_ok=True)
    files = {}

    if category_hierarchy:
        wb = render_category_xlsx.invoke({"paths": category_hierarchy, "blank_path": blank_category_path})
        wb.save(f"output/{fingerprint}_category.xlsx")
        wb.close()
        files["category"] = f"output/{fingerprint}_category.xlsx"

    wb = render_attribute_xlsx.invoke({"defs": attribute_definitions, "blank_path": blank_attribute_path})
    wb.save(f"output/{fingerprint}_attribute.xlsx")
    wb.close()
    files["attribute"] = f"output/{fingerprint}_attribute.xlsx"

    if reference_values:
        wb = render_reference_xlsx.invoke({"refs": reference_values})
        wb.save(f"output/{fingerprint}_reference.xlsx")
        wb.close()
        files["reference"] = f"output/{fingerprint}_reference.xlsx"

    wb = render_product_xlsx.invoke({"rows": product_rows, "attr_names": attr_names})
    wb.save(f"output/{fingerprint}_product.xlsx")
    wb.close()
    files["product"] = f"output/{fingerprint}_product.xlsx"

    return files


@tool
def render_product_xlsx(rows: list[dict], attr_names: list[str]) -> Workbook:
    """Generate product master workbook.
    
    6 fixed columns (Category Path, Variant Attributes, Parent SKU, Code, sku_name, mrp)
    + dynamic attribute columns + 9 image columns."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Category Path"
    ws["B1"] = "Variant Attributes"
    ws["C1"] = "Parent SKU"
    ws["D1"] = "Code"
    ws["E1"] = "sku_name"
    ws["F1"] = "mrp"

    for i, name in enumerate(attr_names, 7):
        ws.cell(row=1, column=i, value=name)

    for i in range(1, 10):
        ws.cell(row=1, column=6 + len(attr_names) + i, value=f"image_{i}")

    for ri, record in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=record.get("category_path"))
        ws.cell(row=ri, column=4, value=record.get("code"))
        ws.cell(row=ri, column=5, value=record.get("sku_name"))
        ws.cell(row=ri, column=6, value=record.get("mrp"))

        for ci, name in enumerate(attr_names, 7):
            ws.cell(row=ri, column=ci, value=record.get(name))

        for ii in range(1, 10):
            key = f"image_{ii}"
            if key in record:
                ws.cell(row=ri, column=6 + len(attr_names) + ii, value=record[key])

    return wb
