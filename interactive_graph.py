import json
import os
import logging

from dotenv import load_dotenv
from google import genai as google_genai
from langgraph.graph import StateGraph, START, END
from langgraph.checkpoint.memory import MemorySaver
from langgraph.prebuilt import ToolNode, tools_condition
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.tools import tool, InjectedToolCallId
from langchain_core.messages import ToolMessage
from typing import Annotated
from langgraph.prebuilt import InjectedState
from langgraph.types import Command

import re

import openpyxl
import json as _json

from interactive_state import InteractiveIngestionState, PhaseOutput
from helpers import (
    read_file, take_rows, fingerprint_headers, extract_image_columns,
    build_product_rows, download_blank_template,
    load_cached_mapping, save_cached_mapping,
)
from tools.mapping import build_attribute_definitions
from tools.references import extract_reference_values
from tools.rendering import render_all_templates
from tools_enrichment import enrich_descriptions
from tools_merger import merge_duplicates
from tools_sheets import merge_sheets_programmatically
from tools.profiling import profile_columns
from learning import fetch_similar_examples
from state import PIM_DEFAULTS, ColumnMapping

load_dotenv()
logger = logging.getLogger("pim_interactive")
logger.setLevel(logging.INFO)
if not logger.handlers:
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S"))
    logger.addHandler(handler)

api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
client = google_genai.Client(api_key=api_key)


# ─── Helpers ──────────────────────────────────────────────────────

def _llm_json(prompt: str, temperature: float = 1.0) -> dict:
    from rate_limiter import wait_for_capacity, track_cost
    wait_for_capacity()
    resp = client.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=prompt,
        config={
            "response_mime_type": "application/json",
            "temperature": temperature,
        },
    )
    track_cost(prompt, resp.text)
    try:
        return json.loads(resp.text)
    except json.JSONDecodeError:
        logger.warning(f"LLM JSON parse failed, raw:\n{resp.text[:500]}")
        return {}


_VALIDATE_TAXONOMY_PROMPT = """You are a PIM taxonomy expert. Review these extracted category paths.

Paths: {paths}

Determine if these are actual product category classifications (like "Apparel > Men > Shoes")
or if they look like product codes, SKUs, serial numbers, or item names (like "1012B862.001").

Return JSON:
{{
  "is_valid": true/false,
  "reason": "If invalid, explain why and suggest what kind of columns to look for instead."
}}
"""


def _validate_taxonomy_llm(paths: list[str]) -> tuple[bool, str]:
    if not paths:
        return True, ""
    prompt = _VALIDATE_TAXONOMY_PROMPT.format(paths=json.dumps(paths[:20]))
    result = _llm_json(prompt, temperature=0.2)
    is_valid = result.get("is_valid", True)
    reason = result.get("reason", "")
    if not is_valid:
        logger.info(f"taxonomy | invalid | {reason[:80]}")
    return is_valid, reason


def _make_empty_phase() -> PhaseOutput:
    return PhaseOutput(
        explanation="",
        reasoning="",
        suggestions=[],
        approved=False,
        user_feedback="",
    )


def _check_type_compatibility(samples: list[str], declared_type: str) -> list[str]:
    if declared_type in ("varchar", "varchar[]"):
        return []
    non_matching = []
    for v in samples:
        v = v.strip()
        if not v:
            continue
        if declared_type == "int":
            try:
                int(v)
            except ValueError:
                non_matching.append(v)
        elif declared_type == "float":
            cleaned = v.replace(",", "").replace("$", "").replace("€", "").replace("£", "").replace("₹", "")
            try:
                float(cleaned)
            except ValueError:
                non_matching.append(v)
        elif declared_type == "boolean":
            if v.lower() not in ("true", "false", "yes", "no", "0", "1", "y", "n", "t", "f"):
                non_matching.append(v)
        elif declared_type == "date":
            date_patterns = [
                r"^\d{4}-\d{2}-\d{2}$",
                r"^\d{2}/\d{2}/\d{4}$",
                r"^\d{2}-\d{2}-\d{4}$",
                r"^\d{4}/\d{2}/\d{2}$",
                r"^\d{2}\.\d{2}\.\d{4}$",
            ]
            if not any(re.match(p, v) for p in date_patterns):
                non_matching.append(v)
    return non_matching


def _validate_mappings(headers: list[str], mapping_data: list[dict], sample_rows_data: list[list]) -> list[dict]:
    errors = []
    col_index = {h: i for i, h in enumerate(headers)}
    for item in mapping_data:
        src = item.get("column", "")
        declared_type = item.get("data_type", "varchar")
        target = item.get("mapped_to", "")
        if not src or not target:
            continue
        if src not in col_index:
            errors.append({"field": target, "issue": f"source_column '{src}' not found in headers", "samples": []})
            continue
        idx = col_index[src]
        samples = []
        for row in sample_rows_data:
            if idx < len(row) and row[idx] is not None and str(row[idx]).strip():
                samples.append(str(row[idx]).strip())
        if not samples:
            continue
        bad = _check_type_compatibility(samples, declared_type)
        if bad and len(bad) / max(len(samples), 1) >= 0.2:
            errors.append({"field": target, "issue": f"Type mismatch: declared '{declared_type}' but samples don't conform", "samples": bad[:5]})
    mapped_targets = {m.get("mapped_to", "") for m in mapping_data}
    for default in PIM_DEFAULTS:
        if default not in mapped_targets:
            errors.append({"field": default, "issue": f"Missing mandatory PIM default: '{default}' has no mapping", "samples": []})
    return errors


# ─── Triage Node ─────────────────────────────────────────────────

def triage_interactive(state: InteractiveIngestionState) -> dict:
    """Open file, detect sheets/headers, seed profile_data and greeting."""
    # Skip if already profiled (e.g. on re-invoke from respond handler)
    if state.get("profile_data") is not None and state.get("profile_data", {}).get("headers"):
        logger.debug("triage | already profiled — skipping")
        return {}

    path = state["file_path"]
    ext = os.path.splitext(path)[1].lower()
    sheet_name = state.get("sheet_name")

    sheets = []
    sheet_count = 0
    first_rows = []
    total_row_count = 0

    if ext == ".csv":
        sheet_count = 1
        gen = read_file(path)
        first_rows = take_rows(gen, 20)
        total_row_count = 1 + sum(1 for _ in gen)
    elif ext in (".xlsx", ".xls"):
        import openpyxl
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            sheet_count = len(sheets)
            if sheet_name and sheet_name in sheets:
                ws = wb[sheet_name]
            else:
                sheet_name = sheets[0]
                ws = wb[sheet_name]
            total_row_count = ws.max_row or 0
            wb.close()
            gen = read_file(path, sheet_name)
            first_rows = take_rows(gen, 20)
        except Exception:
            import xlrd
            xl = xlrd.open_workbook(path)
            sheets = xl.sheet_names()
            sheet_count = len(sheets)
            if not sheet_name or sheet_name not in sheets:
                sheet_name = sheets[0]
            ws = xl.sheet_by_index(sheets.index(sheet_name))
            total_row_count = ws.nrows
            gen = read_file(path, sheet_name)
            first_rows = take_rows(gen, 20)
    else:
        gen = read_file(path, sheet_name or None)
        first_rows = take_rows(gen, 20)
        sheet_count = 1
        total_row_count = 1 + sum(1 for _ in gen)

    # Detect header row via LLM (handles metadata rows between header and data)
    from agents import detect_header_via_llm
    header_row, data_start_row = detect_header_via_llm(first_rows)

    headers = [str(c) if c is not None else "" for c in first_rows[header_row]]
    if data_start_row < header_row + 1:
        data_start_row = header_row + 1
    row_count = total_row_count - data_start_row
    if row_count < 0:
        row_count = 0
    column_count = len(headers)

    state["sheet_name"] = sheet_name
    state["profile_data"] = {
        "headers": headers,
        "sample_rows": [],
        "row_count": row_count,
        "column_count": column_count,
        "header_row": header_row,
        "data_start_row": data_start_row,
    }

    # ── Multi-sheet: collect all sheet metadata ───────────── 
    state["all_sheets"] = [] 
    if ext in (".xlsx", ".xls"): 
        try: 
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True) 
            for sn in sheets: 
                ws = wb[sn] 
                sheet_headers = None
                for r in ws.iter_rows(max_row=5, values_only=True):
                    vals = [str(c).strip() if c else "" for c in r]
                    if any(v for v in vals if v):
                        sheet_headers = vals
                        break
                first_row = sheet_headers if sheet_headers else [] 
                sh = [str(c).strip() if c else "" for c in first_row] 
                state["all_sheets"].append({ 
                    "name": sn, 
                    "headers": [h for h in sh if h], 
                    "row_count": ws.max_row or 0, 
                }) 
            wb.close() 
        except Exception: 
            pass 

    state["phases_completed"] = []
    state["current_phase"] = "categories"
    state["categories"] = _make_empty_phase()
    state["attributes"] = _make_empty_phase()
    state["references"] = _make_empty_phase()
    state["products"] = _make_empty_phase()
    state["core_mappings"] = {}
    state["custom_mappings"] = {}
    state["mapping_confidence"] = {}
    state["generated_files"] = []
    state["remaining_steps"] = 0

    # Build a greeting message
    greeting = (
        f"I've loaded <strong>{os.path.basename(path)}</strong> "
        f"({row_count} rows, {column_count} columns on sheet "
        f"'{sheet_name}').\n\n"
        f"Here is the plan we will work through 4 steps together:\n\n"
        f"1. <strong>Categories</strong> — I will discover your product hierarchy\n"
        f"2. <strong>Attributes</strong> — I will map your columns to PIM fields\n"
        f"3. <strong>Reference Masters</strong> — I will extract dropdown values\n"
        f"4. <strong>Products</strong> — I will compile the final template\n\n"
        f"Ready to start with <strong>Categories</strong>?"
    )
    logger.info(f"triage | file={path} | rows={row_count} | cols={column_count}")

    # Return only deltas — reducer appends to existing state without duplication
    return {
        "messages": [{"role": "assistant", "content": greeting}],
        "profile_data": {
            "headers": headers,
            "sample_rows": [],
            "row_count": row_count,
            "column_count": column_count,
            "header_row": header_row,
            "data_start_row": data_start_row,
        },
        "sheet_name": sheet_name,
        "all_sheets": state.get("all_sheets", []),
        "categories": _make_empty_phase(),
        "attributes": _make_empty_phase(),
        "references": _make_empty_phase(),
        "products": _make_empty_phase(),
        "current_phase": "categories",
        "phases_completed": [],
        "core_mappings": {},
        "custom_mappings": {},
        "mapping_confidence": {},
        "generated_files": [],
        "remaining_steps": 0,
    }


# ─── Categories Phase Node ───────────────────────────────────────

# The actual category logic lives in agents.py (resolve_category_paths).
# This node delegates to it, then wraps the result in a PhaseOutput.


def answer_category_query(query: str, paths: list, explanation: str) -> str:
    prompt = f"""
Answer the user's question based ONLY on the category data below.
If the answer isn't in the data, say so politely.

Discovered category paths: {json.dumps(paths[:30])}
Previous analysis: {explanation[:400] if explanation else '(none)'}
User question: "{query}"

Return JSON: {{"answer": "your response here"}}
"""
    result = _llm_json(prompt)
    return result.get("answer", "I don't have that information about your categories.")


def parse_category_feedback(feedback: str, headers: list | None = None) -> dict:
    header_hint = ""
    if headers:
        header_hint = "Actual column names in the file: " + ", ".join(headers[:30])
    prompt = f"""
Analyze this user feedback for product category onboarding. 

If the user is asking about something NOT related to PIM/product categories (e.g. general chat, jokes, weather, programming help), 
set is_off_topic = true and provide a polite redirect. 

If the user is responding to a merge question (saying yes/no to linking sheets), set is_merge_approval or is_merge_rejection. 

Otherwise, determine if they are explicitly asking to combine or build the category tree from specific columns.
Match the user's column names to the actual headers listed below. Return the EXACT header names as they appear in the file.
{header_hint} 

User feedback: "{feedback}" 

Return valid JSON: 
{{ 
    "is_off_topic": false, 
    "is_merge_approval": false, 
    "is_merge_rejection": false, 
    "is_direct_override": false, 
    "specified_columns": [], 
    "redirect_message": "", 
    "explanation": "" 
}} 

JSON: 
""" 

    try:
        return _llm_json(prompt)
    except Exception: 
        return {"is_off_topic": False, "is_merge_approval": False, "is_merge_rejection": False, 
                "is_direct_override": False, "specified_columns": [], "redirect_message": "", "explanation": ""} 



def build_paths_from_generator(file_path: str, sheet_name: str | None, columns: list[str]) -> list[str]:
    gen = read_file(file_path, sheet_name)
    headers_row = None
    for row in gen:
        vals = [c for c in row if c is not None and str(c).strip() and str(c).strip().lower() != "none"]
        if len(vals) >= 10:
            headers_row = row
            break
    if headers_row is None:
        return []
    headers_raw = [str(c).strip() if c is not None else "" for c in headers_row]
    col_indices = []
    for col in columns:
        col_clean = col.strip()
        for i, h in enumerate(headers_raw):
            if h == col_clean or h.lower() == col_clean.lower():
                col_indices.append(i)
                break
    if not col_indices:
        return []
    paths = set()
    for row in gen:
        parts = []
        for idx in col_indices:
            if idx < len(row):
                val = row[idx]
                if val is not None and str(val).strip().lower() != "nan" and str(val).strip() != "":
                    parts.append(str(val).strip())
        if parts:
            paths.add(" > ".join(parts))
    return sorted(list(paths))

def execute_sheet_merge(file_path: str, base_sheet: str, join_sheet: str, key_col: str) -> str | None: 
    """Merge two sheets by key column, write result to uploads/merged_{uuid}.xlsx""" 
    import openpyxl 
    from openpyxl import Workbook 
    try: 
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True) 
        # Read base sheet 
        ws_base = wb[base_sheet] 
        base_headers = []
        for r in ws_base.iter_rows(max_row=5, values_only=True):
            vals = [str(c).strip() if c else "" for c in r]
            if any(v for v in vals if v):
                base_headers = vals
                break 
        if key_col not in base_headers: 
            return None 
        key_idx = base_headers.index(key_col) 
        base_rows = {} 
        for row in ws_base.iter_rows(values_only=True): 
            if row[key_idx] is not None: 
                base_rows[str(row[key_idx]).strip()] = list(row) 
        # Read join sheet 
        ws_join = wb[join_sheet] 
        join_headers = []
        for r in ws_join.iter_rows(max_row=5, values_only=True):
            vals = [str(c).strip() if c else "" for c in r]
            if any(v for v in vals if v):
                join_headers = vals
                break 
        if key_col not in join_headers: 
            return None 
        join_key_idx = join_headers.index(key_col) 
        join_cols = [c for c in join_headers if c != key_col] 
        result_headers = base_headers + join_cols 
        # Write merged 
        out = Workbook() 
        ws_out = out.active 
        ws_out.append(result_headers) 
        for row in ws_join.iter_rows(values_only=True): 
            key = str(row[join_key_idx]).strip() if row[join_key_idx] else "" 
            if key in base_rows: 
                base_row = base_rows[key] 
                join_vals = [row[join_headers.index(c)] if c in join_headers and join_headers.index(c) < len(row) else "" for c in join_cols] 
                ws_out.append(list(base_row) + join_vals) 
        wb.close() 
        os.makedirs("uploads", exist_ok=True) 
        import uuid 
        out_path = f"uploads/merged_{uuid.uuid4().hex}.xlsx" 
        out.save(out_path) 
        out.close() 
        return out_path 
    except Exception as exc: 
        logger.warning(f"merge failed: {exc}") 
        return None 



def categories_phase(state: InteractiveIngestionState) -> dict:
    file_path = state["file_path"]
    sheet_name = state.get("sheet_name")

    categories_state = state.get("categories", {})
    feedback = categories_state.get("user_feedback", "").strip()
    
         # ── Multi-sheet merge detection ───────────────────────── 
    all_sheets = state.get("all_sheets", []) 
    sheet_merge = state.get("sheet_merge", {}) 
    
    if len(all_sheets) >= 2 and not sheet_merge.get("user_responded") and not feedback: 
        merge_prompt = f""" 
This file has {len(all_sheets)} sheets with product data. 
Here are the sheets and their column headers: 
{json.dumps(all_sheets, indent=2)} 

Can any two sheets be linked together using a common key column  
(e.g. SKU, Product ID, Code, Item Code)? 
Return JSON: 
{{ 
    "can_merge": false, 
    "reasoning": "", 
    "base_sheet": "", 
    "join_sheet": "", 
    "key_column": "", 
    "user_message": "" 
}} 
""" 
        merge_result = _llm_json(merge_prompt, temperature=0.3)
        state["sheet_merge"] = merge_result
        if merge_result.get("can_merge") and not sheet_merge.get("user_responded"):
            state["sheet_merge"]["pending_message"] = merge_result.get("user_message", "") 


    # ── Intent parsing bypass ────────────────────────────────
    merged_path = None
    if feedback:
        headers_for_parse = state.get("profile_data", {}).get("headers", [])
        decision = parse_category_feedback(feedback, headers_for_parse)
        if decision.get("is_merge_approval") or decision.get("is_merge_rejection"):
            state["sheet_merge"]["user_responded"] = True
            if decision.get("is_merge_approval"):
                merge = state.get("sheet_merge", {})
                merged_path = execute_sheet_merge(
                    file_path,
                    merge.get("base_sheet", ""),
                    merge.get("join_sheet", ""),
                    merge.get("key_column", ""),
                )
                if merged_path:
                    state["file_path"] = merged_path
                    msg = f"Merged sheets using '{merge.get('key_column')}' key. Proceeding with the merged data."
                else:
                    msg = "Could not merge sheets. Proceeding with the primary sheet only."
            else:
                msg = "No problem — I'll proceed with the primary sheet."
            state.setdefault("messages", []).append({"role": "assistant", "content": msg})
            return state
        if decision.get("is_off_topic"):
            redirect = decision.get("redirect_message", "Let's keep the focus on your product data onboarding.")
            state.setdefault("messages", []).append({"role": "assistant", "content": redirect})
            logger.info(f"categories | off-topic | redirect sent")
            return state
        if decision.get("is_direct_override") and decision.get("specified_columns"):
            updated_paths = build_paths_from_generator(file_path, sheet_name, decision["specified_columns"])
            if updated_paths:
                explanation = (
                    f"I processed your request: '{decision['explanation']}'. "
                    f"Reconstructed taxonomy from columns: {', '.join(decision['specified_columns'])}."
                )
                suggestions = [
                    {"type": "item", "label": p, "confidence": 100, "reasoning": "Built from your specified columns"}
                    for p in updated_paths
                ]
                state["categories"] = PhaseOutput(
                    explanation=explanation,
                    reasoning=f"Programmatic extraction from columns: {decision['specified_columns']}",
                    suggestions=suggestions,
                    approved=True,
                    user_feedback="",
                )
                state["profile_data"]["category_hierarchy"] = updated_paths
                if len(updated_paths) > 10:
                    vis = "\n".join(f"- {p}" for p in updated_paths[:5])
                    hid = "\n".join(f"- {p}" for p in updated_paths[5:])
                    pd = vis + f"\n<details><summary>+{len(updated_paths)-5} more paths</summary>\n{hid}\n</details>"
                else:
                    pd = "\n".join(f"- {p}" for p in updated_paths)
                msg = f"<strong>Category Discovery</strong>\n\n{explanation}\n\n{pd}\n\nFound <strong>{len(updated_paths)}</strong> paths."
                state.setdefault("messages", []).append({"role": "assistant", "content": msg})
                logger.info(f"categories | bypass | paths={len(updated_paths)} | cols={decision['specified_columns']}")
                return state

    # ── Conversational fallback ───────────────────────────────
    if feedback:
        existing = state.get("categories", {})
        existing_paths = state.get("profile_data", {}).get("category_hierarchy", [])
        existing_explanation = existing.get("explanation", "")
        if existing_explanation or existing_paths:
            answer = answer_category_query(feedback, existing_paths, existing_explanation)
            state.setdefault("messages", []).append({"role": "assistant", "content": answer})
            logger.info(f"categories | conversational | answered from state")
            return state

    # ── Standard extraction via agents.py ────────────────────
    profile = state.get("profile_data", {}) or {}
    headers = profile.get("headers", [])

    cat_state = {
        "source_path": file_path,
        "sheet_name": sheet_name,
        "headers": headers,
        "header_row": profile.get("header_row", 0),
        "data_start_row": profile.get("data_start_row", 1),
        "category_candidates": [],
        "category_path_config": {},
        "category_hierarchy": [],
        "need_user_input": False,
        "is_known_schema": False,
        "mapping": [],
        "sample_rows": [],
        "row_count": profile.get("row_count", 0),
    }

    from agents import resolve_category_paths
    resolve_category_paths(cat_state)

    paths = cat_state.get("category_hierarchy", [])
    explanation = cat_state.get("category_reasoning", "")
    needs_input = cat_state.get("need_user_input", False)

    suggestions = [
        {"type": "item", "label": p, "confidence": 95, "reasoning": "Part of the product category hierarchy"}
        for p in paths
    ]
    if len(paths) > 10:
        visible = "\n".join(f"- {p}" for p in paths[:5])
        hidden = "\n".join(f"- {p}" for p in paths[5:])
        path_display = visible + f"\n<details><summary>+{len(paths)-5} more paths</summary>\n{hidden}\n</details>"
    else:
        path_display = "\n".join(f"- {p}" for p in paths)

    if not explanation:
        if paths:
            explanation = f"I discovered <strong>{len(paths)}</strong> category paths from your data."
        else:
            explanation = "I wasn't able to automatically detect a clear category hierarchy."

    merge_pending = state.get("sheet_merge", {}).get("pending_message", "")
    if merge_pending:
        explanation = merge_pending + "\n\n" + explanation

    state["categories"] = PhaseOutput(
        explanation=explanation,
        reasoning=f"Strategy: declarative recipe on {len(paths)} paths.",
        suggestions=suggestions,
        approved=False,
        user_feedback=feedback,
    )
    state["profile_data"]["category_hierarchy"] = paths

    msg = (
        f"<strong>Category Discovery</strong>\n\n{explanation}"
        + (f"\n\n{path_display}\n\nFound <strong>{len(paths)}</strong> paths. Do these look right?" if paths else "")
    )
    state.setdefault("messages", []).append({"role": "assistant", "content": msg})
    logger.info(f"categories | paths={len(paths)} | need_input={needs_input}")
    return state


# ─── Attributes Phase Node ───────────────────────────────────────

SCREENING_PROMPT = """You are a VinAI PIM onboarding assistant screening a data file for attribute discovery.

Current phase: Attribute Screening (Step 1 of 2)

File: {filename}

Column names:
{column_names}

Sample data (first {sample_count} rows):
{samples}

Metadata rows found above the header (constraints, descriptions, types):
{metadata_context}

Entity-Attribute-Value format detection:
{eav_info}

Your job is to screen the columns and answer:
1. What format is the data? ("columns" = standard, "eav" = attribute names in rows, "hybrid", "unknown")
2. Which columns are actual product attributes vs metadata/internal fields vs noise?
3. Exclude columns that are: internal IDs, audit timestamps, system flags, empty/constant columns, purely descriptive metadata
4. Report your sampling confidence. If 50 rows isn't enough variety to decide, set needs_more_samples to true.

Return JSON:
{{
  "format_detected": "columns",
  "attribute_columns": ["Product Name", "Brand", "Price", ...],
  "excluded_columns": {{"Internal Code": "internal ID — not a product attribute", "Created At": "system timestamp"}},
  "sampling_confidence": 90,
  "needs_more_samples": false,
  "reasoning": "Brief explanation of format and screening decisions."
}}
"""


MAPPING_PROMPT = """You are a VinAI PIM onboarding assistant mapping screened columns to PIM attributes.

Current phase: Attribute Mapping (Step 2 of 2)

These columns have been confirmed as product attributes. Map each one.

File: {filename}

Column profiles (unique counts + samples) for screened attributes:
{profiles}

Sample data (first rows):
{samples}

Historical corrections for similar columns:
{few_shots}

The user's feedback from the previous attempt (if any):
{feedback}

Previous validation errors (fix these):
{validation_errors_text}

Include BOTH `attribute_type` AND `attribute_data_type` for every mapping. Group your results into three buckets:

1. <strong>High-Confidence Core Mappings</strong> — system-critical fields (sku_name, code, mrp)
   that are clearly identified.
2. <strong>Custom Dynamic Attributes</strong> — proprietary columns that should be preserved.
3. <strong>Low-Confidence / Ambiguous Fields</strong> — columns where you're < 80% sure.

PIM defaults (map TO these, don't recreate): sku_name, code, mrp

attribute_type rules:
- Brand, colour, size, gender, season, type, category → Dropdown (constraint=true)
- Description/notes → RichText (length=65536)
- Codes, names, numbers, prices → Textbox
- Multi-value tags/features → MultiSelect (constraint=true)
- Multi-value predefined choices → MultiSelectDropdown (constraint=true)
- Text with predefined options + free-text entry → MultiTextBox (constraint=true)
- Date fields → Date
- Image URLs → Textbox, length=2048

attribute_data_type rules:
- Prices, decimals → float
- Counts, quantities → int
- Yes/No fields → boolean
- Dates → date
- Everything else → varchar

Group each custom attribute into a logical PIM group:
- weight, length, width, height, material, size → Shipping & Dimensions
- brand, manufacturer, country_of_origin, coo → Brand & Origin
- processor, ram, storage, battery, screen → Technical Specs
- heel_height, sole, closure_type, fit → Sizing & Fit
- meta_title, meta_description, url_key → SEO
- color, pattern, fabric, care_instructions → Product Details
- Default → Basic Information

Return JSON:
{{
  "explanation": "A plain-English summary of the mappings.",
  "reasoning": "Technical details.",
  "suggestions": [
    {{
      "type": "group",
      "label": "High-Confidence Core Mappings",
      "items": [{{"type": "item", "column": "Product Name", "mapped_to": "sku_name", "attribute_type": "Textbox", "attribute_data_type": "varchar", "attribute_group": "Basic Information", "confidence": 100, "reasoning": "..."}}]
    }},
    {{"type": "group", "label": "Custom Dynamic Attributes", "items": [...]}},
    {{"type": "group", "label": "Low-Confidence / Needs Review", "items": [...]}}
  ]
}}
"""


def _detect_eav_format(headers: list, rows: list) -> dict:
    if not rows or len(rows) < 3 or len(headers) < 2:
        return {"is_eav": False, "attr_column": None, "val_column": None}
    col_value_counts = {}
    for ci in range(min(3, len(headers))):
        vals = [str(r[ci]).strip() for r in rows[:50] if ci < len(r) and r[ci] is not None and str(r[ci]).strip()]
        unique = set(vals)
        if len(vals) >= 5 and 3 <= len(unique) <= 100 and (len(unique) / len(vals)) < 0.7:
            col_value_counts[ci] = len(unique)
    if not col_value_counts:
        return {"is_eav": False, "attr_column": None, "val_column": None}
    likely_attr_col = max(col_value_counts, key=col_value_counts.get)
    likely_val_col = 1 if likely_attr_col == 0 else 0
    attr_samples = list(set(str(r[likely_attr_col]).strip() for r in rows[:50] if likely_attr_col < len(r) and r[likely_attr_col] is not None))
    return {
        "is_eav": True,
        "attr_column": headers[likely_attr_col] if likely_attr_col < len(headers) else "",
        "val_column": headers[likely_val_col] if likely_val_col < len(headers) else "",
        "sample_attributes": sorted(attr_samples)[:15],
    }


def _read_metadata_rows(state) -> str:
    profile = state.get("profile_data", {}) or {}
    hr = profile.get("header_row", 0)
    dr = profile.get("data_start_row", hr + 1)
    if dr <= hr + 1:
        return "(no metadata rows — headers are immediately above data)"
    headers = profile.get("headers", [])
    gen = read_file(state["file_path"], state.get("sheet_name"))
    all_rows = list(gen)
    metadata_rows = []
    for mr in range(hr + 1, min(dr, len(all_rows))):
        row_data = {}
        for ci, h in enumerate(headers):
            if ci < len(all_rows[mr]) and all_rows[mr][ci] is not None and str(all_rows[mr][ci]).strip():
                row_data[h] = str(all_rows[mr][ci]).strip()[:80]
        if row_data:
            metadata_rows.append(row_data)
    if metadata_rows:
        return json.dumps(metadata_rows[:5], indent=2)
    return "(no metadata rows found)"


def parse_attribute_feedback(feedback: str) -> dict:
    prompt = f"""
Analyze this user feedback about PIM attribute mappings.
Determine if they are asking to add, remove, or change column-to-attribute mappings.

If the user is asking about something NOT related to PIM/product attributes (e.g. general chat, jokes, weather, programming help),
set is_off_topic = true and provide a polite redirect.

User feedback: "{feedback}"

Return valid JSON:
{{
    "is_off_topic": false,
    "has_override": false,
    "add_mappings": [],
    "remove_columns": [],
    "remap": [],
    "redirect_message": "",
    "explanation": ""
}}

Rules:
- add_mappings: when user says "map X to Y" and X is a source column
- remove_columns: when user says "remove/drop/skip X"
- remap: when user says "change X to map to Y instead"
- type can be: Textbox, Dropdown, RichText, Textarea, MultiSelect, Date
- data_type can be: varchar, int, float, boolean, date

JSON:
"""
    try:
        return _llm_json(prompt)
    except Exception:
        return {"is_off_topic": False, "has_override": False, "add_mappings": [], "remove_columns": [], "remap": [], "redirect_message": "", "explanation": ""}


def attributes_phase(state: InteractiveIngestionState) -> dict:
    profile = state.get("profile_data", {}) or {}
    headers = profile.get("headers", [])
    feedback = state.get("attributes", {}).get("user_feedback", "")
    cycle = state.get("attributes", {}).get("correction_cycle", 0)

    gen = read_file(state["file_path"], state.get("sheet_name"))
    hr = profile.get("header_row", 0)
    dr = profile.get("data_start_row", hr + 1)
    for _ in range(dr):
        try:
            next(gen)
        except StopIteration:
            break
    all_data_rows = list(gen)

    fp = fingerprint_headers(headers)
    cached = load_cached_mapping(fp)
    if cached and not feedback and cycle == 0:
        return _apply_cached_mappings(state, cached, headers, fp)

    # ── Intent parsing bypass ────────────────────────────────
    if feedback:
        decision = parse_attribute_feedback(feedback)
        if decision.get("is_off_topic"):
            redirect = decision.get("redirect_message", "Let's keep the focus on your product data attributes.")
            state.setdefault("messages", []).append({"role": "assistant", "content": redirect})
            logger.info(f"attributes | off-topic | redirect sent")
            return state
        if decision.get("has_override"):
            core = dict(state.get("core_mappings", {}))
            custom = dict(state.get("custom_mappings", {}))
            for col in decision.get("remove_columns", []):
                core = {k: v for k, v in core.items() if v != col}
                custom.pop(col, None)
            for r in decision.get("remap", []):
                col = r.get("column", "")
                new_tgt = r.get("new_target", "")
                for k, v in list(core.items()):
                    if v == col:
                        core[new_tgt] = core.pop(k)
                        break
                if col in custom:
                    custom[new_tgt] = custom.pop(col)
            for m in decision.get("add_mappings", []):
                col = m.get("column", "")
                tgt = m.get("mapped_to", col)
                if tgt in ("sku_name", "code", "mrp"):
                    core[tgt] = col
                elif col and col in headers:
                    custom[col] = tgt
            state["core_mappings"] = core
            state["custom_mappings"] = custom
            all_items = []
            for tgt, col in core.items():
                all_items.append({"column": col, "mapped_to": tgt, "attribute_type": "Textbox", "attribute_data_type": "varchar", "attribute_group": "Basic Information", "confidence": 100})
            for col, tgt in custom.items():
                all_items.append({"column": col, "mapped_to": tgt, "attribute_type": "Textbox", "attribute_data_type": "varchar", "attribute_group": "Basic Information", "confidence": 100})
            state["attributes"] = PhaseOutput(
                explanation=decision.get("explanation", "Applied your changes."),
                reasoning="Bypass: user-specified mapping overrides.",
                suggestions=[
                    {"type": "group", "label": "High-Confidence Core Mappings", "items": [
                        {"type": "item", "column": col, "mapped_to": tgt, "attribute_type": "Textbox", "attribute_data_type": "varchar", "attribute_group": "Basic Information", "confidence": 100, "reasoning": "User override"}
                        for tgt, col in core.items()
                    ]},
                    {"type": "group", "label": "Custom Dynamic Attributes", "items": [
                        {"type": "item", "column": col, "mapped_to": tgt, "attribute_type": "Textbox", "attribute_data_type": "varchar", "attribute_group": "Basic Information", "confidence": 100, "reasoning": "User override"}
                        for col, tgt in custom.items()
                    ]},
                ],
                approved=True,
                user_feedback="",
            )
            msg = f"<strong>Attribute Mapping</strong>\n\n{decision.get('explanation', 'Applied your changes.')}"
            state.setdefault("messages", []).append({"role": "assistant", "content": msg})
            logger.info(f"attributes | bypass | core={len(core)} custom={len(custom)}")
            return state

    # ── Step 1: Column screening (adaptive) ───────────────────
    max_sample = min(len(all_data_rows), 500)
    sample_size = min(50, max_sample)
    sampling_round = 0
    max_sampling_rounds = 3
    needs_more = True
    attr_columns = []

    while needs_more and sampling_round < max_sampling_rounds and sample_size <= max_sample:
        sampling_round += 1
        current_rows = all_data_rows[:sample_size]

        sample_rows = current_rows[:5]
        samples = []
        for row in sample_rows:
            s = {}
            for i, h in enumerate(headers):
                if i < len(row) and row[i] is not None and str(row[i]).strip():
                    s[h] = str(row[i]).strip()[:80]
            samples.append(s)

        eav_info = _detect_eav_format(headers, current_rows)
        eav_text = json.dumps(eav_info, indent=2) if eav_info["is_eav"] else "(standard columnar format)"
        metadata_context = _read_metadata_rows(state)
        col_names_text = "\n".join(f"{i+1}. {h}" for i, h in enumerate(headers[:60]))

        screen_prompt = SCREENING_PROMPT.format(
            filename=os.path.basename(state["file_path"]),
            column_names=col_names_text,
            samples=json.dumps(samples, indent=2),
            sample_count=sample_size,
            metadata_context=metadata_context,
            eav_info=eav_text,
        )

        screen_result = _llm_json(screen_prompt)
        attr_columns = screen_result.get("attribute_columns", [])
        needs_more = screen_result.get("needs_more_samples", False)

        if needs_more:
            sample_size = min(sample_size * 2, max_sample)
            logger.info(f"attributes | screen round {sampling_round}: {sample_size} rows, {len(attr_columns)} attr cols")

    # If screening found nothing useful, fall back to all headers
    if not attr_columns:
        attr_columns = [h for h in headers if h.strip()]
        logger.info("attributes | screening returned empty — using all headers")

    # ── Step 2: Full mapping on screened columns ──────────────
    cols = profile_columns.invoke({"headers": headers, "rows": all_data_rows})
    attr_col_set = set(attr_columns)
    filtered_profiles = [
        {"name": c["name"], "unique": c["unique"], "non_null": c["non_null"], "sample": c["sample"][:3]}
        for c in cols if c["name"] in attr_col_set
    ]

    sample_rows = all_data_rows[:5]
    samples = []
    for row in sample_rows:
        s = {}
        for i, h in enumerate(headers):
            if i < len(row) and row[i] is not None and str(row[i]).strip():
                s[h] = str(row[i]).strip()[:80]
        samples.append(s)

    few_shots = []
    seen_targets = set()
    for h in attr_columns[:10]:
        vals = []
        for row in all_data_rows:
            idx = headers.index(h)
            if idx < len(row) and row[idx] is not None and str(row[idx]).strip():
                vals.append(str(row[idx]).strip()[:40])
        if not vals:
            continue
        matches = fetch_similar_examples(h, vals, k=2)
        for m in matches:
            tgt = m["target_attribute"]
            if tgt and tgt not in seen_targets:
                few_shots.append(m)
                seen_targets.add(tgt)
                if len(few_shots) >= 5:
                    break
        if len(few_shots) >= 5:
            break

    few_shots_text = "\n".join(
        f'- "{fs["column_name"]}" → {fs["target_attribute"]} ({fs["attribute_type"]}, {fs["attribute_data_type"]}, mandatory={str(fs["mandatory"]).lower()})'
        for fs in few_shots
    ) if few_shots else "(no historical corrections available)"

    previous_errors = state.get("attributes", {}).get("validation_errors", [])
    validation_errors_text = "\n".join(
        f'- {e["field"]}: {e["issue"]}'
        for e in previous_errors
    ) if previous_errors else "(none)"

    map_prompt = MAPPING_PROMPT.format(
        filename=os.path.basename(state["file_path"]),
        profiles=json.dumps(filtered_profiles[:40], indent=2),
        samples=json.dumps(samples, indent=2),
        few_shots=few_shots_text,
        feedback=feedback or "(none — first attempt)",
        validation_errors_text=validation_errors_text,
    )

    result = _llm_json(map_prompt)

    all_items = []
    for group in result.get("suggestions", []):
        if group.get("type") == "group":
            for item in group.get("items", []):
                if item.get("type") == "item" and item.get("column") and item.get("mapped_to"):
                    all_items.append(item)

    validation_errors = _validate_mappings(headers, all_items, all_data_rows[:10])

    max_retries = 3
    if validation_errors and cycle < max_retries and not feedback:
        new_cycle = cycle + 1
        state.setdefault("attributes", {})["correction_cycle"] = new_cycle
        state.setdefault("attributes", {})["validation_errors"] = validation_errors
        logger.info(f"attributes | auto-retry {new_cycle}/{max_retries} | errors={len(validation_errors)}")
        return attributes_phase(state)

    cache_data = [{"source_column": it["column"], "target_attribute": it["mapped_to"],
                    "attribute_type": it.get("attribute_type", "Textbox"),
                    "attribute_data_type": it.get("attribute_data_type", "varchar")}
                   for it in all_items]
    save_cached_mapping(fp, cache_data)

    core_group = {}
    custom_group = {}
    for group in result.get("suggestions", []):
        if group.get("type") == "group":
            label = group.get("label", "")
            for item in group.get("items", []):
                if item.get("type") == "item":
                    col = item.get("column", "")
                    mapped = item.get("mapped_to", "")
                    if label == "High-Confidence Core Mappings" and mapped:
                        core_group[mapped] = col
                    elif label == "Custom Dynamic Attributes":
                        custom_group[col] = col

    state["core_mappings"] = core_group
    state["custom_mappings"] = custom_group

    explanation = result.get("explanation", "")
    if validation_errors:
        error_text = "; ".join(f'{e["field"]}: {e["issue"]}' for e in validation_errors[:3])
        explanation += f"\n\n\u26a0\ufe0f <strong>{len(validation_errors)} validation issue(s)</strong>: {error_text}"

    state["attributes"] = PhaseOutput(
        explanation=explanation,
        reasoning=result.get("reasoning", ""),
        suggestions=result.get("suggestions", []),
        approved=False,
        user_feedback=feedback,
    )
    if validation_errors:
        state["attributes"]["validation_errors"] = validation_errors

    msg = (
        f"\ud83d\udccb <strong>Attribute Mapping</strong>\n\n{explanation}\n\n"
        f"I've grouped the mappings below. You can accept all, or tell me "
        f"about specific ones you'd like to change."
    )
    state.setdefault("messages", []).append({"role": "assistant", "content": msg})

    logger.info(f"attributes | screened {len(attr_columns)} cols | core={len(core_group)} custom={len(custom_group)} errors={len(validation_errors)}")
    return state


def _apply_cached_mappings(state, cached, headers, fp):
    core_group = {}
    custom_group = {}
    suggestions = []
    core_items = []
    custom_items = []
    for m in cached:
        tgt = m.get("target_attribute", "")
        src = m.get("source_column", "")
        if tgt in ("sku_name", "code", "mrp"):
            core_group[tgt] = src
            core_items.append({"type": "item", "column": src, "mapped_to": tgt,
                               "attribute_type": m.get("attribute_type", "Textbox"),
                               "attribute_data_type": m.get("attribute_data_type", "varchar"),
                               "attribute_group": m.get("attribute_group", "Basic Information"),
                               "confidence": 100, "reasoning": "Cached from previous session"})
        else:
            custom_group[src] = src
            custom_items.append({"type": "item", "column": src, "mapped_to": tgt,
                                 "attribute_type": m.get("attribute_type", "Textbox"),
                                 "attribute_data_type": m.get("attribute_data_type", "varchar"),
                                 "attribute_group": m.get("attribute_group", "Basic Information"),
                                 "confidence": 100, "reasoning": "Cached from previous session"})
    if core_items:
        suggestions.append({"type": "group", "label": "High-Confidence Core Mappings", "items": core_items})
    if custom_items:
        suggestions.append({"type": "group", "label": "Custom Dynamic Attributes", "items": custom_items})
    state["core_mappings"] = core_group
    state["custom_mappings"] = custom_group
    state["attributes"] = PhaseOutput(
        explanation=f"Loaded {len(cached)} mappings from cache (fingerprint: {fp}).",
        reasoning="",
        suggestions=suggestions,
        approved=False,
        user_feedback="",
    )
    msg = (
        f"\ud83d\udccb <strong>Attribute Mapping</strong>\n\nI recognized this file structure — "
        f"I've loaded <strong>{len(cached)}</strong> saved mappings from a previous session.\n\n"
        f"You can accept all, or tell me about specific ones you'd like to change."
    )
    state.setdefault("messages", []).append({"role": "assistant", "content": msg})
    logger.info(f"attributes | cache hit | fp={fp} | mappings={len(cached)}")
    return state


# ─── References Phase Node ──────────────────────────────────────

REFERENCES_PROMPT = """You are a VinAI PIM onboarding assistant. The user has confirmed their attribute mappings.

Current phase: Reference Masters

Here are the reference values I extracted from the data:
{refs_summary}

The user's feedback from the previous attempt (if any):
{feedback}

Attributes marked as Dropdown or MultiSelect in a PIM need a strict,
predefined list of allowed options — called Reference Masters. This
prevents data-entry mistakes and typos.

Review the extracted reference masters above and:
1. Generate an educational explanation of what Reference Masters are and why they matter
2. Flag any messy/inconsistent values (e.g. "MED" vs "M", "Blk" vs "Black")
3. Suggest normalizations for messy values

Return JSON:
{{
  "explanation": "An educational paragraph explaining what Reference Masters are and why they matter.",
  "reasoning": "Details about normalizations proposed.",
  "suggestions": [
    {{
      "type": "item",
      "label": "Brand",
      "column": "Brand Name",
      "unique_count": 15,
      "values": ["Sony", "Bose", "Samsung"],
      "messy_values": [],
      "normalizations": [],
      "confidence": 100
    }}
  ]
}}
"""


def parse_reference_feedback(feedback: str) -> dict:
    prompt = f"""
Analyze this user feedback about reference master values.
Determine if they are approving, or if they want to override specific normalizations.

If the user is asking about something NOT related to PIM/reference data (e.g. general chat, jokes, weather),
set is_off_topic = true and provide a polite redirect.

User feedback: "{feedback}"

Return valid JSON:
{{
    "is_off_topic": false,
    "is_approval": false,
    "keep_values": {{"value_to_keep": "reason"}},
    "override_normalizations": [{{"original": "MED", "keep_as": "MED"}}],
    "redirect_message": "",
    "explanation": ""
}}

Rules:
- is_approval: True if user says yes/looks good/proceed/approve
- keep_values: dict of values the user explicitly wants to keep as-is (e.g. {{"SML": "keep as SML"}})
- override_normalizations: list of specific normalization overrides

JSON:
"""
    try:
        return _llm_json(prompt)
    except Exception:
        return {"is_off_topic": False, "is_approval": False, "keep_values": {}, "override_normalizations": [], "redirect_message": "", "explanation": ""}


def references_phase(state: InteractiveIngestionState) -> dict:
    profile = state.get("profile_data", {}) or {}
    headers = profile.get("headers", [])
    feedback = state.get("references", {}).get("user_feedback", "")

    gen = read_file(state["file_path"], state.get("sheet_name"))
    hr = profile.get("header_row", 0)
    dr = profile.get("data_start_row", hr + 1)
    for _ in range(dr):
        try:
            next(gen)
        except StopIteration:
            break
    all_data_rows = list(gen)

    # Profile all columns (gets unique_values for reference extraction)
    cols = profile_columns.invoke({"headers": headers, "rows": all_data_rows})

    # Use programmatic extract_reference_values with the mappings we have
    mapping_dicts = []
    for target, col in state.get("core_mappings", {}).items():
        if col:
            mapping_dicts.append({"source_column": col, "target_attribute": target, "attribute_type": "Textbox"})
    for col, preserved in state.get("custom_mappings", {}).items():
        mapping_dicts.append({"source_column": col, "target_attribute": preserved, "attribute_type": "Dropdown"})

    refs = extract_reference_values.invoke({"mappings": mapping_dicts, "profiles": cols})

    # Build suggestions from programmatic refs
    suggestions = []
    for master_name, values in refs.items():
        suggestions.append({
            "type": "item",
            "label": master_name.replace(" Master", ""),
            "column": master_name,
            "unique_count": len(values),
            "values": values[:20],
            "messy_values": [],
            "normalizations": [],
            "confidence": 100,
        })

    # ── Intent parsing bypass ────────────────────────────────
    if feedback:
        decision = parse_reference_feedback(feedback)
        if decision.get("is_off_topic"):
            redirect = decision.get("redirect_message", "Let's keep the focus on your reference data.")
            state.setdefault("messages", []).append({"role": "assistant", "content": redirect})
            logger.info("references | off-topic | redirect sent")
            return state
        if decision.get("is_approval"):
            state["references"] = PhaseOutput(
                explanation="Reference masters approved.",
                reasoning="User confirmed reference values.",
                suggestions=suggestions,
                approved=True,
                user_feedback="",
            )
            state.setdefault("messages", []).append({"role": "assistant", "content": "Reference masters approved. Moving on..."})
            logger.info("references | bypass | approved")
            return state
        if decision.get("keep_values"):
            keep = decision.get("keep_values", {})
            for s in suggestions:
                s["messy_values"] = [v for v in s.get("messy_values", []) if v not in keep]
                s["normalizations"] = [n for n in s.get("normalizations", [])
                                       if not (isinstance(n, dict) and n.get("original") in keep)
                                       and not (isinstance(n, str) and n.split(" → ")[0].strip() in keep)]
            state["references"] = PhaseOutput(
                explanation=decision.get("explanation", "Applied your reference overrides."),
                reasoning="Bypass: user-specified normalization overrides.",
                suggestions=suggestions,
                approved=True,
                user_feedback="",
            )
            state.setdefault("messages", []).append({"role": "assistant", "content": f"Applied your overrides. Moving on..."})
            logger.info("references | bypass | keep overrides applied")
            return state

    # Use LLM to add educational explanation + detect messy values
    refs_summary = "\n".join(
        f'- {k}: {len(v)} values — {v[:5]}'
        for k, v in refs.items()
    ) if refs else "(no reference masters detected)"

    prompt = REFERENCES_PROMPT.format(
        refs_summary=refs_summary,
        feedback=feedback or "(none — first attempt)",
    )
    result = _llm_json(prompt)

    # Merge LLM's messy value detection into programmatic suggestions
    llm_suggestions = result.get("suggestions", [])
    for llm_s in llm_suggestions:
        llm_label = llm_s.get("label", "")
        for s in suggestions:
            if s["label"] == llm_label or s["column"] == llm_s.get("column", ""):
                if llm_s.get("messy_values"):
                    s["messy_values"] = llm_s["messy_values"]
                if llm_s.get("normalizations"):
                    s["normalizations"] = llm_s["normalizations"]
                break

    state["references"] = PhaseOutput(
        explanation=result.get("explanation", ""),
        reasoning=result.get("reasoning", ""),
        suggestions=suggestions,
        approved=False,
        user_feedback=feedback,
    )

    msg = (
        f"\ud83d\udcda <strong>Reference Masters</strong>\n\n{result.get('explanation', '')}\n\n"
        f"I extracted <strong>{len(suggestions)}</strong> reference lists from your data. "
        f"Let me know if any values need cleaning up!"
    )
    state.setdefault("messages", []).append({"role": "assistant", "content": msg})

    logger.info(f"references | masters={len(suggestions)}")
    return state


# ─── Products Phase Node ────────────────────────────────────────

PRODUCTS_PROMPT = """You are a VinAI PIM onboarding assistant. The user has confirmed their
Reference Masters.

Current phase: Product Preview

Here is a preview of the first 3 mapped products:
{preview}

The user's feedback from the previous attempt (if any):
{feedback}

The product sheet will have:
- Fixed columns: Category Path, Variant Attributes, Parent SKU, Code, sku_name, mrp
- Dynamic columns: one per attribute from the mapping phase
- Image URL columns: image_1 through image_9

Explain to the user what they're seeing and confirm they're happy to proceed.

Return JSON:
{{
  "explanation": "A friendly summary of what the product sheet will look like, referencing the preview.",
  "reasoning": "Technical breakdown of row count, columns, and image handling.",
  "suggestions": []
}}
"""


def parse_product_feedback(feedback: str) -> dict:
    prompt = f"""
Analyze this user feedback about product compilation.
Determine if they want to exclude any columns from the output, or if they're approving.

If the user is asking about something NOT related to PIM/product compilation (e.g. general chat, jokes, weather, programming help),
set is_off_topic = true and provide a polite redirect.

User feedback: "{feedback}"

Return valid JSON:
{{
    "is_off_topic": false,
    "has_override": false,
    "exclude_columns": [],
    "is_approval": false,
    "redirect_message": "",
    "explanation": ""
}}

Rules:
- has_override: True if user wants to exclude/remove/drop specific columns
- exclude_columns: list of column names to exclude (based on source column names or attribute names)
- is_approval: True if user is saying yes/looks good/proceed/go ahead

JSON:
"""
    try:
        return _llm_json(prompt)
    except Exception:
        return {"is_off_topic": False, "has_override": False, "exclude_columns": [], "is_approval": False, "redirect_message": "", "explanation": ""}


def products_phase(state: InteractiveIngestionState) -> dict:
    profile = state.get("profile_data", {}) or {}
    headers = profile.get("headers", [])
    row_count = profile.get("row_count", 0)
    column_count = profile.get("column_count", 0)
    feedback = state.get("products", {}).get("user_feedback", "")

    gen = read_file(state["file_path"], state.get("sheet_name"))
    hr = profile.get("header_row", 0)
    dr = profile.get("data_start_row", hr + 1)
    for _ in range(dr):
        try:
            next(gen)
        except StopIteration:
            break
    all_data_rows = list(gen)

    # Build product rows programmatically (same as render does)
    row_mappings = []
    for target, col in state.get("core_mappings", {}).items():
        if col:
            row_mappings.append({"source_column": col, "target_attribute": target})
    for col, preserved in state.get("custom_mappings", {}).items():
        row_mappings.append({"source_column": col, "target_attribute": preserved})

    # ── Intent parsing bypass ────────────────────────────────
    if feedback:
        decision = parse_product_feedback(feedback)
        if decision.get("is_off_topic"):
            redirect = decision.get("redirect_message", "Let's keep the focus on your product data compilation.")
            state.setdefault("messages", []).append({"role": "assistant", "content": redirect})
            logger.info(f"products | off-topic | redirect sent")
            return state
        if decision.get("is_approval"):
            state["products"] = PhaseOutput(
                explanation="Proceeding with product compilation.",
                reasoning="User approved.",
                suggestions=[],
                approved=True,
                user_feedback="",
            )
            state.setdefault("messages", []).append({"role": "assistant", "content": "Generating final PIM templates..."})
            logger.info("products | bypass | approved")
            return state
        if decision.get("has_override"):
            excluded = set(decision.get("exclude_columns", []))
            filtered_mappings = [m for m in row_mappings
                                 if m.get("source_column") not in excluded
                                 and m.get("target_attribute") not in excluded]
            if filtered_mappings:
                row_mappings = filtered_mappings

    img_cols = extract_image_columns(headers, row_mappings)
    product_rows = build_product_rows(
        headers, all_data_rows, row_mappings, img_cols, state.get("core_mappings"),
    )

    # ── Image URL validation ─────────────────────────────────
    img_url_count = 0
    img_broken = 0
    img_samples = []
    for pr in product_rows:
        for ii in range(1, 10):
            url = pr.get(f"image_{ii}", "")
            if url and isinstance(url, str) and url.strip():
                img_url_count += 1
                url = url.strip()
                if not url.startswith("http"):
                    img_broken += 1
                    if len(img_samples) < 3:
                        img_samples.append(url)
    img_warning = ""
    if img_url_count > 0 and (img_broken / max(img_url_count, 1)) > 0.3:
        pct = int(100 * img_broken / img_url_count)
        img_warning = (
            f"\n\n<strong>{pct}% of image links appear invalid</strong> "
            f"(e.g. '{img_samples[0] if img_samples else ''}'). "
            f"Proceed anyway?"
        )

    # Build a preview of the first 3 mapped products
    preview_rows = []
    for pr in product_rows[:3]:
        row_preview = {k: str(v)[:40] for k, v in list(pr.items())[:8]}
        preview_rows.append(row_preview)

    attr_count = len(state.get("core_mappings", {})) + len(state.get("custom_mappings", {}))
    img_col_count = len(img_cols[:9])
    total_cols = 6 + attr_count + img_col_count

    preview_text = json.dumps(preview_rows, indent=2) if preview_rows else "(no product rows)"

    prompt = PRODUCTS_PROMPT.format(
        preview=preview_text,
        feedback=feedback or "(none — first attempt)",
    )
    result = _llm_json(prompt)

    suggestions = [
        {"type": "item", "label": "Total products", "value": str(row_count),
         "reasoning": f"Found {row_count} data rows in the source file"},
        {"type": "item", "label": "Total columns in output", "value": str(total_cols),
         "reasoning": f"6 fixed columns + {attr_count} attributes + {img_col_count} image columns"},
        {"type": "item", "label": "Image columns detected", "value": str(img_col_count),
         "reasoning": "Up to 9 images per product supported"},
    ]
    if preview_rows:
        suggestions.insert(0, {
            "type": "group",
            "label": "Product Preview (first 3 rows)",
            "items": [
                {"type": "item", "column": k, "mapped_to": v, "confidence": 100,
                 "reasoning": ""}
                for pr in preview_rows for k, v in pr.items()
            ][:12],
        })

    explanation = (result.get("explanation", "") + img_warning).strip()

    state["products"] = PhaseOutput(
        explanation=explanation,
        reasoning=result.get("reasoning", ""),
        suggestions=suggestions,
        approved=False,
        user_feedback=feedback,
    )

    msg = (
        f"\ud83d\udce6 <strong>Product Compilation</strong>\n\n{result.get('explanation', '')}\n\n"
        f"<strong>{row_count}</strong> products across <strong>{total_cols}</strong> columns ready.\n\n"
        f"Shall I proceed with generating the final PIM templates?"
    )
    state.setdefault("messages", []).append({"role": "assistant", "content": msg})

    logger.info(f"products | rows={row_count} | cols={total_cols}")
    return state


# ─── Render Node ────────────────────────────────────────────────

def render_interactive(state: InteractiveIngestionState) -> dict:
    """Generate the 4 output xlsx files after all phases are approved."""
    fp = fingerprint_headers(state.get("profile_data", {}).get("headers", []))
    headers = state.get("profile_data", {}).get("headers", [])
    cats = state.get("profile_data", {}).get("category_hierarchy", [])

    # Build mapping objects from phase outputs
    from state import ColumnMapping

    mapping_objs = []
    for target, col in state.get("core_mappings", {}).items():
        if col:
            mapping_objs.append(ColumnMapping(
                source_column=col, target_attribute=target, confidence=1.0,
            ))
    for col, preserved in state.get("custom_mappings", {}).items():
        mapping_objs.append(ColumnMapping(
            source_column=col, target_attribute=preserved, confidence=1.0,
        ))

    # Build attribute definitions
    attr_defs = build_attribute_definitions.invoke({"mappings": mapping_objs})

    # Extract reference values
    mapping_dicts = [
        {"source_column": m.source_column, "target_attribute": m.target_attribute,
         "attribute_type": m.attribute_type}
        for m in mapping_objs
    ]
    profiles_list = state.get("profile_data", {}).get("profiles", [])
    refs = extract_reference_values.invoke({
        "mappings": mapping_dicts,
        "profiles": profiles_list,
    })

    # Read all data rows
    rows = read_file(state["file_path"], state.get("sheet_name"))
    hr = state.get("profile_data", {}).get("header_row", 0)
    dr = max(state.get("profile_data", {}).get("data_start_row", hr + 1), hr + 1)
    for _ in range(dr):
        try:
            next(rows)
        except StopIteration:
            break

    # Build product rows
    img_cols = extract_image_columns(headers, mapping_dicts)
    row_mappings = [{"source_column": m.source_column, "target_attribute": m.target_attribute} for m in mapping_objs]
    product_rows = build_product_rows(
        headers, rows, row_mappings, img_cols, state.get("core_mappings"),
    )

    # Download blank templates if JWT available
    jwt = state.get("jwt_token", "")
    blank_cat = download_blank_template(jwt, "category") if jwt else ""
    blank_attr = download_blank_template(jwt, "attribute") if jwt else ""

    # Render
    attr_names = [m.target_attribute for m in mapping_objs]
    files = render_all_templates.invoke({
        "fingerprint": fp,
        "category_hierarchy": cats,
        "attribute_definitions": attr_defs,
        "reference_values": refs,
        "headers": headers,
        "product_rows": product_rows,
        "attr_names": attr_names,
        "blank_category_path": blank_cat,
        "blank_attribute_path": blank_attr,
    })

    state["generated_files"] = list(files.values())
    state["product_rows"] = product_rows
    state["current_phase"] = "complete"

    msg = (
        f"<strong>All done!</strong> I've generated <strong>{len(files)}</strong> PIM template files:\n\n"
        + "\n".join(f"- `{v.split('/')[-1]}`" for v in files.values())
        + "\n\nYou can download them now. They're ready for upload to your PIM."
    )
    state.setdefault("messages", []).append({"role": "assistant", "content": msg})

    logger.info(f"render | files={list(files.values())}")
    return state


# ═══════════════════════════════════════════════════════════════
# AGENT TOOLS — Called by LLM via ToolNode
# Each tool:
#   - Accepts only LLM-facing args (file path, sheet name, column hints)
#   - Reads state via InjectedState for programmatic context
#   - Returns a string result for the LLM to read
#   - Writes structured data directly to state
# ═══════════════════════════════════════════════════════════════

from typing import Annotated
from langgraph.prebuilt import InjectedState


@tool
def profile_file(
    file_path: str,
    sheet_name: str | None = None,
    state: Annotated[dict, InjectedState()] = None,
    tool_call_id: Annotated[str, InjectedToolCallId] = "",
) -> Command:
    """Analyse a spreadsheet structure. Call first on new files.
    Detects headers, rows, columns, and cached mappings.

    Args:
        file_path: Path to CSV/XLSX/XLS file.
        sheet_name: Sheet name (auto-detected if omitted).
    """
    import json as _json
    # Resolve file path — LLM may pass bare filename missing uploads/ prefix
    if not os.path.exists(file_path):
        state_path = state.get("file_path", "") if state else ""
        if state_path and os.path.exists(state_path):
            file_path = state_path
        else:
            trial = os.path.join("uploads", os.path.basename(file_path))
            if os.path.exists(trial):
                file_path = trial

    ext = os.path.splitext(file_path)[1].lower()
    # Detect header row & sheet
    if not sheet_name:
        if ext in (".xlsx", ".xls"):
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_name = wb.sheetnames[0]
            wb.close()
        else:
            sheet_name = "auto-detect"

    gen = read_file(file_path, sheet_name)
    first_rows = take_rows(gen, 20)

    from agents import detect_header_via_llm
    header_row, data_start_row = detect_header_via_llm(first_rows)

    headers = [str(c) if c is not None else "" for c in first_rows[header_row]]
    if data_start_row < header_row + 1:
        data_start_row = header_row + 1

    # Count rows
    row_count = len(first_rows)  # partial — full count from existing triage if available
    try:
        all_rows = list(read_file(file_path, sheet_name))
        row_count = len(all_rows) - data_start_row
        if row_count < 0:
            row_count = 0
    except Exception:
        pass

    # Profile columns
    data_rows = list(read_file(file_path, sheet_name)) if row_count > 20 else all_rows
    hr = header_row
    dr = max(data_start_row, hr + 1)
    data_rows_list = list(data_rows) if not isinstance(data_rows, list) else data_rows
    try:
        gen2 = read_file(file_path, sheet_name)
        for _ in range(dr):
            next(gen2, None)
        data_rows_list = list(gen2)
    except Exception:
        data_rows_list = []

    cols = []
    for ci, h in enumerate(headers):
        vals = []
        for row in data_rows_list:
            if ci < len(row) and row[ci] is not None and str(row[ci]).strip():
                vals.append(str(row[ci]))
        uniq = list(set(vals))
        sample = (uniq[:3] + uniq[-2:]) if len(uniq) > 5 else uniq
        cols.append({
            "name": h,
            "non_null": len(vals),
            "unique": len(uniq),
            "sample": sample,
            "unique_values": uniq if len(uniq) <= 100 else [],
        })

    fp = fingerprint_headers(headers)
    cached = load_cached_mapping(fp)

    # Collect multi-sheet metadata
    all_sheets = []
    if ext in (".xlsx", ".xls"):
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                sh = [str(c).strip() if c else "" for c in next(ws.iter_rows(max_row=1, values_only=True), [])]
                all_sheets.append({"name": sn, "headers": [h for h in sh if h], "row_count": ws.max_row or 0})
            wb.close()
        except Exception:
            pass

    # Write to state
    state["sheet_name"] = sheet_name
    state["profile_data"] = {
        "headers": headers,
        "sample_rows": [{headers[i]: str(r[i])[:60] for i in range(min(len(headers), len(r))) if r[i] is not None and str(r[i]).strip()} for r in data_rows_list[:3]],
        "row_count": row_count,
        "column_count": len(headers),
        "header_row": header_row,
        "data_start_row": data_start_row,
        "profiles": cols,
        "category_hierarchy": [],
    }
    state["all_sheets"] = all_sheets

    # Initialize empty phase outputs
    state["categories"] = _make_empty_phase()
    state["attributes"] = _make_empty_phase()
    state["references"] = _make_empty_phase()
    state["products"] = _make_empty_phase()
    state["core_mappings"] = {}
    state["custom_mappings"] = {}
    state["mapping_confidence"] = {}
    state["generated_files"] = []
    state["product_rows"] = []

    # Mark phase complete
    state.setdefault("completed_phases", [])
    if "triage" not in state["completed_phases"]:
        state["completed_phases"].append("triage")

    cache_note = f" I also found {len(cached)} saved mappings from a previous session." if cached else ""
    return Command(
        update={
            "sheet_name": sheet_name,
            "profile_data": {
                "headers": headers,
                "sample_rows": [{headers[i]: str(r[i])[:60] for i in range(min(len(headers), len(r))) if r[i] is not None and str(r[i]).strip()} for r in data_rows_list[:3]],
                "row_count": row_count,
                "column_count": len(headers),
                "header_row": header_row,
                "data_start_row": data_start_row,
                "profiles": cols,
                "category_hierarchy": [],
            },
            "all_sheets": all_sheets,
            "categories": _make_empty_phase(),
            "attributes": _make_empty_phase(),
            "references": _make_empty_phase(),
            "products": _make_empty_phase(),
            "core_mappings": {},
            "custom_mappings": {},
            "mapping_confidence": {},
            "generated_files": [],
            "product_rows": [],
            "completed_phases": ["triage"],
            "messages": [
                ToolMessage(
                    content=(
                        f"Profiled **{os.path.basename(file_path)}** — sheet **{sheet_name}**, "
                        f"{row_count} rows, {len(headers)} columns.{cache_note}\n\n"
                        f"Headers: {', '.join(headers[:10])}{'...' if len(headers) > 10 else ''}"
                    ),
                    tool_call_id=tool_call_id,
                )
            ],
        },
    )


@tool
def extract_categories(
    file_path: str,
    sheet_name: str | None = None,
    specified_columns: list[str] | None = None,
    state: Annotated[dict, InjectedState()] = None,
    tool_call_id: Annotated[str, InjectedToolCallId] = "",
) -> Command:
    """Extract product category hierarchy from spreadsheet columns.
    Uses 5-strategy fallback. Pass specified_columns if you know which cols.

    Args:
        file_path: Path to source file.
        sheet_name: Sheet name (omit for profiled sheet).
        specified_columns: Column names to join as category path.
    """
    # Resolve file path — LLM may pass bare filename missing uploads/ prefix
    if not os.path.exists(file_path):
        state_path = state.get("file_path", "")
        if state_path and os.path.exists(state_path):
            file_path = state_path
        else:
            trial = os.path.join("uploads", os.path.basename(file_path))
            if os.path.exists(trial):
                file_path = trial

    if not sheet_name:
        sheet_name = state.get("sheet_name")

    profile = state.get("profile_data", {}) or {}
    headers = profile.get("headers", [])

    state.setdefault("completed_phases", [])
    if "categories" not in state["completed_phases"]:
        state["completed_phases"].append("categories")

    if specified_columns:
        # Direct reconstruction from user-specified columns
        updated_paths = build_paths_from_generator(file_path, sheet_name, specified_columns)
        if updated_paths:
            is_valid, reason = _validate_taxonomy_llm(updated_paths)
            if not is_valid:
                return Command(update={"messages": [ToolMessage(content=f"The extracted values look like product codes rather than categories. {reason}", tool_call_id=tool_call_id)]})
            state["profile_data"]["category_hierarchy"] = updated_paths
            state["profile_data"]["category_hierarchy"] = updated_paths
            state["categories"] = PhaseOutput(
                explanation=f"Built from columns: {', '.join(specified_columns)}",
                reasoning=f"Programmatic extraction from {len(specified_columns)} columns",
                suggestions=[{"type": "item", "label": p, "confidence": 100, "reasoning": "Direct column extraction"} for p in updated_paths],
                approved=True,
                user_feedback="",
            )
            path_summary = "\n".join(f"- {p}" for p in updated_paths[:8])
            return Command(
                update={
                    "profile_data": {
                        **state.get("profile_data", {}),
                        "category_hierarchy": updated_paths,
                    },
                    "categories": PhaseOutput(
                        explanation=f"Built from columns: {', '.join(specified_columns)}",
                        reasoning=f"Programmatic extraction from {len(specified_columns)} columns",
                        suggestions=[{"type": "item", "label": p, "confidence": 100, "reasoning": "Direct column extraction"} for p in updated_paths],
                        approved=True,
                        user_feedback="",
                    ),
                    "messages": [ToolMessage(content=(
                        f"Extracted **{len(updated_paths)}** category paths from columns "
                        f"{', '.join(specified_columns)}.\n\n"
                        f"{path_summary}"
                        + (f"\n\n+{len(updated_paths) - 8} more paths" if len(updated_paths) > 8 else "")
                    ), tool_call_id=tool_call_id)]
                },
            )
        return Command(update={"messages": [ToolMessage(content="Could not extract paths from those columns. Try different column names.", tool_call_id=tool_call_id)]})

    # Standard 5-strategy fallback
    cat_state = {
        "source_path": file_path,
        "sheet_name": sheet_name,
        "headers": headers,
        "header_row": profile.get("header_row", 0),
        "data_start_row": profile.get("data_start_row", 1),
        "category_candidates": [],
        "category_path_config": {},
        "category_hierarchy": [],
        "need_user_input": False,
        "is_known_schema": False,
        "mapping": [],
        "sample_rows": [],
        "row_count": profile.get("row_count", 0),
    }

    from agents import resolve_category_paths
    resolve_category_paths(cat_state)

    paths = cat_state.get("category_hierarchy", [])
    if paths:
        is_valid, reason = _validate_taxonomy_llm(paths)
        if not is_valid:
            return Command(update={"messages": [ToolMessage(content=f"The extracted values look like product codes rather than categories. {reason}", tool_call_id=tool_call_id)]})
            
    explanation = cat_state.get("category_reasoning", "")
    needs_input = cat_state.get("need_user_input", False)

    cat_update = {
        "profile_data": {
            **state.get("profile_data", {}),
            "category_hierarchy": paths,
        },
        "categories": PhaseOutput(
            explanation=explanation or f"Discovered {len(paths)} category paths.",
            reasoning=f"Strategy fallback chain on {len(paths)} paths.",
            suggestions=[{"type": "item", "label": p, "confidence": 95, "reasoning": "Part of product hierarchy"} for p in paths],
            approved=not needs_input,
            user_feedback="",
        ),
    }

    if not paths:
        cat_update["messages"] = [ToolMessage(content="I wasn't able to detect a clear category hierarchy from this file.", tool_call_id=tool_call_id)]
        return Command(update=cat_update)

    path_summary = "\n".join(f"- {p}" for p in paths[:8])
    cat_update["messages"] = [ToolMessage(content=(
        f"Discovered **{len(paths)}** category paths.\n\n{path_summary}"
        + (f"\n\n+{len(paths) - 8} more paths" if len(paths) > 8 else "")
        + ("\n\nSome paths may need review — do they look correct?" if needs_input else "")
    ), tool_call_id=tool_call_id)]
    
    return Command(update=cat_update)


@tool
def map_attributes(
    file_path: str,
    sheet_name: str | None = None,
    feedback: str | None = None,
    state: Annotated[dict, InjectedState()] = None,
    tool_call_id: Annotated[str, InjectedToolCallId] = "",
) -> Command:
    """Map source columns to PIM attributes. Core (sku/code/mrp) + custom.
    Screens cols in rounds, validates types, caches by fingerprint.

    Args:
        file_path: Path to source file.
        sheet_name: Sheet name (omit for profiled sheet).
        feedback: User correction text.
    """
    # Resolve file path — LLM may pass bare filename missing uploads/ prefix
    if not os.path.exists(file_path):
        state_path = state.get("file_path", "")
        if state_path and os.path.exists(state_path):
            file_path = state_path
        else:
            trial = os.path.join("uploads", os.path.basename(file_path))
            if os.path.exists(trial):
                file_path = trial

    if not sheet_name:
        sheet_name = state.get("sheet_name")

    profile = state.get("profile_data", {}) or {}
    headers = profile.get("headers", [])

    # Read data
    gen = read_file(file_path, sheet_name)
    hr = profile.get("header_row", 0)
    dr = profile.get("data_start_row", hr + 1)
    for _ in range(dr):
        next(gen, None)
    all_data_rows = list(gen)

    fp = fingerprint_headers(headers)
    cached = load_cached_mapping(fp)
    state.setdefault("completed_phases", [])
    if cached and not feedback:
        state["core_mappings"] = {}
        state["custom_mappings"] = {}
        for m in cached:
            tgt = m.get("target_attribute", "")
            src = m.get("source_column", "")
            if tgt in ("sku_name", "code", "mrp"):
                state["core_mappings"][tgt] = src
            else:
                state["custom_mappings"][src] = src
        completed = list(state.get("completed_phases", []))
        if "attributes" not in completed:
            completed.append("attributes")
        return Command(
            update={
                "core_mappings": state.get("core_mappings", {}),
                "custom_mappings": state.get("custom_mappings", {}),
                "attributes": PhaseOutput(
                    explanation=f"Loaded {len(cached)} mappings from cache.",
                    reasoning="Fingerprint cache hit.",
                    suggestions=[],
                    approved=False,
                    user_feedback="",
                ),
                "completed_phases": completed,
                "messages": [ToolMessage(content=(f"Loaded {len(cached)} saved mappings from previous session (fp : {fp})"), tool_call_id=tool_call_id)]
            },
            # value=f"Loaded **{len(cached)}** saved mappings from a previous session (fingerprint: {fp}).",
        )

    # Step 1: Column screening (adaptive sampling)
    max_sample = min(len(all_data_rows), 500)
    sample_size = min(50, max_sample)
    sampling_round = 0
    max_sampling_rounds = 3
    needs_more = True
    attr_columns = []

    while needs_more and sampling_round < max_sampling_rounds and sample_size <= max_sample:
        sampling_round += 1
        current_rows = all_data_rows[:sample_size]
        sample_rows = current_rows[:5]
        samples = [{h: str(r[i]).strip()[:80] for i, h in enumerate(headers) if i < len(r) and r[i] is not None and str(r[i]).strip()} for r in sample_rows]

        col_names_text = "\n".join(f"{i+1}. {h}" for i, h in enumerate(headers[:60]))
        screen_prompt = SCREENING_PROMPT.format(
            filename=os.path.basename(file_path),
            column_names=col_names_text,
            samples=_json.dumps(samples, indent=2),
            sample_count=sample_size,
            metadata_context="",
            eav_info="(standard columnar format)",
        )
        screen_result = _llm_json(screen_prompt)
        attr_columns = screen_result.get("attribute_columns", [])
        needs_more = screen_result.get("needs_more_samples", False)
        if needs_more:
            sample_size = min(sample_size * 2, max_sample)

    if not attr_columns:
        attr_columns = [h for h in headers if h.strip()]

    # Step 2: Profile columns via Polars (memory-safe for large files)
    from helpers_data_plane import profile_large_file
    profile_result = profile_large_file(file_path, sheet_name)
    all_profiles = {p["column_name"]: p for p in profile_result["profiles"]}
    filtered_profiles = [
        {
            "name": col,
            "unique": all_profiles[col]["unique_count"],
            "non_null": profile_result["row_count"] - all_profiles[col]["null_count"],
            "sample": all_profiles[col]["sample_values"][:3],
        }
        for col in attr_columns if col in all_profiles
    ]

    sample_rows = all_data_rows[:5]
    samples = [{h: str(r[i]).strip()[:80] for i, h in enumerate(headers) if i < len(r) and r[i] is not None and str(r[i]).strip()} for r in sample_rows]

    few_shots_text = "(no historical corrections available)"
    if not feedback:
        few_shots = []
        seen = set()
        for h in attr_columns[:10]:
            vals = [str(r[headers.index(h)]).strip()[:40] for r in all_data_rows if headers.index(h) < len(r) and r[headers.index(h)] is not None and str(r[headers.index(h)]).strip()]
            if vals:
                matches = fetch_similar_examples(h, vals, k=2)
                for m in matches:
                    tgt = m["target_attribute"]
                    if tgt and tgt not in seen:
                        few_shots.append(m)
                        seen.add(tgt)
                        if len(few_shots) >= 5:
                            break
            if len(few_shots) >= 5:
                break
        if few_shots:
            few_shots_text = "\n".join(f'- "{fs["column_name"]}" → {fs["target_attribute"]} ({fs["attribute_type"]})' for fs in few_shots)

    map_prompt = MAPPING_PROMPT.format(
        filename=os.path.basename(file_path),
        profiles=_json.dumps(filtered_profiles[:40], indent=2),
        samples=_json.dumps(samples, indent=2),
        few_shots=few_shots_text,
        feedback=feedback or "(none — first attempt)",
        validation_errors_text="(none)",
    )
    result = _llm_json(map_prompt)

    # Build structured state
    all_items = []
    for group in result.get("suggestions", []):
        if group.get("type") == "group":
            for item in group.get("items", []):
                if item.get("type") == "item" and item.get("column") and item.get("mapped_to"):
                    all_items.append(item)

    core_group = {}
    custom_group = {}
    for group in result.get("suggestions", []):
        if group.get("type") == "group":
            label = group.get("label", "")
            for item in group.get("items", []):
                if item.get("type") == "item":
                    col = item.get("column", "")
                    mapped = item.get("mapped_to", "")
                    if label == "High-Confidence Core Mappings" and mapped:
                        core_group[mapped] = col
                    elif label == "Custom Dynamic Attributes":
                        custom_group[col] = col

    state["core_mappings"] = core_group
    state["custom_mappings"] = custom_group

    # Cache successful mappings
    cache_data = [{"source_column": it["column"], "target_attribute": it["mapped_to"],
                    "attribute_type": it.get("attribute_type", "Textbox"),
                    "attribute_data_type": it.get("attribute_data_type", "varchar")}
                   for it in all_items]
    save_cached_mapping(fp, cache_data)

    completed = list(state.get("completed_phases", []))
    if "attributes" not in completed:
        completed.append("attributes")

    return Command(
        update={
            "core_mappings": core_group,
            "custom_mappings": custom_group,
            "attributes": PhaseOutput(
                explanation=result.get("explanation", "Attribute mapping complete."),
                reasoning=result.get("reasoning", ""),
                suggestions=result.get("suggestions", []),
                approved=False,
                user_feedback="",
            ),
            "completed_phases": completed,
            "messages": [ToolMessage(content=(
                f"Mapped **{len(core_group) + len(custom_group)}** attributes "
                f"({len(core_group)} core + {len(custom_group)} custom).\n\n"
                + ("\n".join(f"- `{col}` → **{tgt}**" for tgt, col in core_group.items())
                   if core_group else "No core mappings identified.")
                + ("\n\nCustom attributes ready for review." if custom_group else "")
            ), tool_call_id=tool_call_id)],
        },
    )


@tool
def extract_references(
    state: Annotated[dict, InjectedState()] = None,
    tool_call_id: Annotated[str, InjectedToolCallId] = "",
) -> Command:
    """Extract unique values for Dropdown/MultiSelect attributes.
    Reads mappings from state. Call after map_attributes.
    No args needed."""
    profile = state.get("profile_data", {}) or {}
    headers = profile.get("headers", [])
    cols = profile.get("profiles", [])

    if not cols and headers:
        # Profile columns via Polars (memory-safe)
        try:
            from helpers_data_plane import profile_large_file
            pr = profile_large_file(state.get("file_path", ""), state.get("sheet_name"))
            cols = []
            for p in pr["profiles"]:
                uniq_vals = []
                try:
                    import polars as pl
                    from helpers_data_plane import get_lazy_frame, detect_header_row_and_headers
                    hr, _, _ = detect_header_row_and_headers(state["file_path"], state.get("sheet_name"))
                    lazy = get_lazy_frame(state["file_path"], state.get("sheet_name"), hr, pr["headers"])
                    uniq_vals = lazy.select(pl.col(p["column_name"]).unique()).collect(streaming=True)[p["column_name"]].to_list()
                    uniq_vals = [str(v) for v in uniq_vals if v is not None]
                except Exception:
                    pass
                cols.append({
                    "name": p["column_name"],
                    "non_null": pr["row_count"] - p["null_count"],
                    "unique": p["unique_count"],
                    "sample": p["sample_values"][:3],
                    "unique_values": uniq_vals[:100] if len(uniq_vals) > 100 else uniq_vals,
                })
        except Exception:
            cols = []

    # Build mapping dicts from state
    mapping_dicts = []
    for target, col in state.get("core_mappings", {}).items():
        if col:
            mapping_dicts.append({"source_column": col, "target_attribute": target, "attribute_type": "Textbox"})
    for col, preserved in state.get("custom_mappings", {}).items():
        mapping_dicts.append({"source_column": col, "target_attribute": preserved, "attribute_type": "Dropdown"})

    refs = extract_reference_values.invoke({"mappings": mapping_dicts, "profiles": cols})

    suggestions = []
    for master_name, values in refs.items():
        suggestions.append({
            "type": "item",
            "label": master_name.replace(" Master", ""),
            "column": master_name,
            "unique_count": len(values),
            "values": values[:20],
            "messy_values": [],
            "normalizations": [],
        })

    state["references"] = PhaseOutput(
        explanation=f"Extracted {len(suggestions)} reference masters.",
        reasoning="Programmatic extraction from column profiles.",
        suggestions=suggestions,
        approved=False,
        user_feedback="",
    )

    completed = list(state.get("completed_phases", []))
    if "references" not in completed:
        completed.append("references")

    ref_update = {
        "references": PhaseOutput(
            explanation=f"Extracted {len(suggestions)} reference masters.",
            reasoning="Programmatic extraction from column profiles.",
            suggestions=suggestions,
            approved=False,
            user_feedback="",
        ),
        "completed_phases": completed,
    }

    if not suggestions:
        ref_update["messages"] = [ToolMessage(content="No Dropdown or MultiSelect attributes found — no reference masters to extract.", tool_call_id=tool_call_id)]
        return Command(update=ref_update)

    summary = "\n".join(f"- **{s['label']}**: {s['unique_count']} values — {', '.join(str(v) for v in s['values'][:5])}" for s in suggestions)
    
    ref_update["messages"] = [ToolMessage(content=f"Extracted **{len(suggestions)}** reference masters.\n\n{summary}", tool_call_id=tool_call_id)]
    
    return Command(update=ref_update)


@tool
def build_products(
    state: Annotated[dict, InjectedState()] = None,
    tool_call_id: Annotated[str, InjectedToolCallId] = "",
) -> Command:
    """Build product rows from confirmed mappings + source data.
    Validates image URLs. Call after map_attributes + extract_references.
    No args needed."""
    profile = state.get("profile_data", {}) or {}
    headers = profile.get("headers", [])
    row_count = profile.get("row_count", 0)

    # Read data rows
    gen = read_file(state.get("file_path", ""), state.get("sheet_name"))
    hr = profile.get("header_row", 0)
    dr = profile.get("data_start_row", hr + 1)
    for _ in range(dr):
        next(gen, None)
    all_data_rows = list(gen)

    # Build row mappings from state
    row_mappings = []
    for target, col in state.get("core_mappings", {}).items():
        if col:
            row_mappings.append({"source_column": col, "target_attribute": target})
    for col, preserved in state.get("custom_mappings", {}).items():
        row_mappings.append({"source_column": col, "target_attribute": preserved})

    img_cols = extract_image_columns(headers, row_mappings)
    product_rows = build_product_rows(
        headers, all_data_rows, row_mappings, img_cols, state.get("core_mappings"),
    )

    # Image URL validation
    img_url_count = 0
    img_broken = 0
    img_samples = []
    for pr in product_rows:
        for ii in range(1, 10):
            url = pr.get(f"image_{ii}", "")
            if url and isinstance(url, str) and url.strip():
                img_url_count += 1
                url = url.strip()
                if not url.startswith("http"):
                    img_broken += 1
                    if len(img_samples) < 3:
                        img_samples.append(url)

    img_warning = ""
    if img_url_count > 0 and (img_broken / max(img_url_count, 1)) > 0.3:
        pct = int(100 * img_broken / img_url_count)
        img_warning = f"\n\n⚠️ **{pct}% of image links appear invalid** (e.g. '{img_samples[0] if img_samples else ''}')."

    attr_count = len(state.get("core_mappings", {})) + len(state.get("custom_mappings", {}))
    total_cols = 6 + attr_count + min(len(img_cols), 9)

    completed = list(state.get("completed_phases", []))
    if "products" not in completed:
        completed.append("products")

    return Command(
        update={
            "product_rows": product_rows if isinstance(product_rows, list) else list(product_rows),
            "products": PhaseOutput(
                explanation=f"Built {len(product_rows)} product rows across {total_cols} columns." + img_warning,
                reasoning=f"{row_count} source rows, {attr_count} attributes, {len(img_cols)} image columns",
                suggestions=[],
                approved=False,
                user_feedback="",
            ),
            "completed_phases": completed,
            "messages": [ToolMessage(content=(
                f"Built **{len(product_rows)}** product rows across **{total_cols}** columns "
                f"({6} fixed + {attr_count} attributes + {min(len(img_cols), 9)} image)."
                + img_warning
            ), tool_call_id=tool_call_id)],
        },
    )


@tool
def render_templates(
    state: Annotated[dict, InjectedState()] = None,
    tool_call_id: Annotated[str, InjectedToolCallId] = "",
) -> Command:
    """Generate 4 PIM xlsx files (category/attribute/reference/product).
    Final step. No args needed."""
    fp = fingerprint_headers(state.get("profile_data", {}).get("headers", []))
    headers = state.get("profile_data", {}).get("headers", [])
    cats = state.get("profile_data", {}).get("category_hierarchy", [])

    from state import ColumnMapping

    # Build mapping objects
    mapping_objs = []
    for target, col in state.get("core_mappings", {}).items():
        if col:
            mapping_objs.append(ColumnMapping(source_column=col, target_attribute=target, confidence=1.0))
    for col, preserved in state.get("custom_mappings", {}).items():
        mapping_objs.append(ColumnMapping(source_column=col, target_attribute=preserved, confidence=1.0))

    if not mapping_objs:
        return Command(update={"messages": [ToolMessage(content="No attribute mappings found. Please run map_attributes first.", tool_call_id=tool_call_id)]},)

    # Build attribute definitions
    attr_defs = build_attribute_definitions.invoke({"mappings": mapping_objs})

    # Extract references from state
    mapping_dicts = [
        {"source_column": m.source_column, "target_attribute": m.target_attribute, "attribute_type": m.attribute_type}
        for m in mapping_objs
    ]
    profiles_list = state.get("profile_data", {}).get("profiles", [])
    refs = extract_reference_values.invoke({"mappings": mapping_dicts, "profiles": profiles_list})

    # Read all data rows
    rows = read_file(state["file_path"], state.get("sheet_name"))
    hr = state.get("profile_data", {}).get("header_row", 0)
    dr = max(state.get("profile_data", {}).get("data_start_row", hr + 1), hr + 1)
    for _ in range(dr):
        next(rows, None)

    img_cols = extract_image_columns(headers, mapping_dicts)
    row_mappings = [{"source_column": m.source_column, "target_attribute": m.target_attribute} for m in mapping_objs]
    from helpers import build_product_rows_streaming
    product_rows = build_product_rows_streaming(
        headers, rows, row_mappings, img_cols, state.get("core_mappings"),
    )

    # Download blank templates if JWT available
    jwt = state.get("jwt_token", "")
    blank_cat = download_blank_template(jwt, "category") if jwt else ""
    blank_attr = download_blank_template(jwt, "attribute") if jwt else ""

    attr_names = [m.target_attribute for m in mapping_objs]
    files = render_all_templates.invoke({
        "fingerprint": fp,
        "category_hierarchy": cats,
        "attribute_definitions": attr_defs,
        "reference_values": refs,
        "headers": headers,
        "product_rows": product_rows,
        "attr_names": attr_names,
        "blank_category_path": blank_cat,
        "blank_attribute_path": blank_attr,
    })

    completed = list(state.get("completed_phases", []))
    if "render" not in completed:
        completed.append("render")

    file_list = "\n".join(f"- `{v.split('/')[-1]}`" for v in files.values())
    return Command(
        update={
            "generated_files": list(files.values()),
            "completed_phases": completed,
            "messages": [ToolMessage(content=(
                f"✅ Generated **{len(files)}** PIM template files:\n\n{file_list}\n\n"
                f"They're saved to the output directory and ready to download."
            ), tool_call_id=tool_call_id)],
        },
    )


# ─── Agent Reason Node ──────────────────────────────────────────

AGENT_SYSTEM_PROMPT = """You are VinGPT, a PIM data onboarding assistant. Guide users from messy spreadsheet to 4 standardized PIM templates.

## Milestones (strict order — check completed_phases before acting)
1. triage (auto): file loaded & profiled
2. categories → extract_categories  | 3. attributes → map_attributes
4. references → extract_references  | 5. products → build_products
6. render → render_templates
7. enrich → enrich_descriptions (if user asks to fill missing descriptions)
8. merge_duplicates → merge_duplicates (if user asks to find/merge near-duplicates)
9. merge_sheets → merge_sheets_programmatically (if file has multiple sheets to join)

Do NOT skip ahead. If a milestone isn't in completed_phases, run it next.

## Rules
- Explain before calling a tool. Present results clearly and ask confirmation.
- If a tool returns empty/fails: do NOT retry the same turn. Tell the user what happened, ask a simple yes/no.
- Don't call enrich_descriptions or merge_duplicates before build_products has run.
- If the workbook has multiple sheets with related data, suggest merge_sheets_programmatically to the user.
- No jargon: use "missing values" not "null", "column layout" not "schema".
- Off-topic user? Redirect back to the current phase politely.
- If the user asks to see/list/show data, present it directly in your response.
- Max 2 tool calls per turn. After 2, stop and present findings."""


def agent_reason_node(state: InteractiveIngestionState) -> dict:
    """Central reasoning engine — decides whether to call a tool or respond conversationally."""
    messages = state["messages"]
    from langchain_core.messages import BaseMessage, SystemMessage, ToolMessage

    # If the last message is from assistant (no pending user input), skip LLM call
    if messages:
        last = messages[-1]
        last_is_assistant = (
            (isinstance(last, BaseMessage) and last.type == "ai") or
            (isinstance(last, dict) and last.get("role") == "assistant")
        )
        if last_is_assistant:
            logger.debug("agent | no pending user input — skipping LLM call")
            return {}  # No delta — state unchanged, router routes to END

    # Windowing: keep last 12 messages, but never split function_call/function_response pairs
    if len(messages) > 12:
        truncated = messages[-12:]
        # Walk forward from the truncation point to find a clean boundary
        # (skip past any ToolMessages and AIMessages with tool_calls)
        for i, msg in enumerate(truncated):
            is_tool = isinstance(msg, ToolMessage)
            is_tool_call = isinstance(msg, BaseMessage) and hasattr(msg, "tool_calls") and msg.tool_calls
            if not is_tool and not is_tool_call:
                messages = truncated[i:]
                break
        else:
            messages = truncated
        logger.info(f"agent | truncated {len(messages)} messages (windowed from larger history)")

    # Convert messages — preserve live LangChain objects, convert plain dicts
    lc_messages = []
    for msg in messages:
        if isinstance(msg, BaseMessage):
            lc_messages.append(msg)
        elif isinstance(msg, dict):
            role = msg.get("role", "user")
            content = msg.get("content", "")
            if role == "user":
                from langchain_core.messages import HumanMessage
                lc_messages.append(HumanMessage(content=content))
            elif role == "assistant":
                from langchain_core.messages import AIMessage
                lc_messages.append(AIMessage(content=content))
            else:
                from langchain_core.messages import HumanMessage
                lc_messages.append(HumanMessage(content=content))

    # Prepend system prompt
    lc_messages = [SystemMessage(content=AGENT_SYSTEM_PROMPT)] + lc_messages

    # Build the agent LLM with tools bound
    agent_tools = [profile_file, extract_categories, map_attributes, extract_references, build_products, render_templates, enrich_descriptions, merge_duplicates, merge_sheets_programmatically]
    llm = ChatGoogleGenerativeAI(
        model="gemini-2.5-flash",
        temperature=1.0,
    ).bind_tools(agent_tools)

    from rate_limiter import wait_for_capacity, track_cost
    wait_for_capacity()
    response = llm.invoke(lc_messages)
    track_cost(str(lc_messages), str(response.content) if hasattr(response, 'content') else '')

    # Bug 2 Fix: Return ONLY the new response (delta).
    return {"messages": [response]}


# ─── Agent Router ───────────────────────────────────────────────

def route_agent_action(state: InteractiveIngestionState) -> str:
    """Pure router — checks for tool_calls and budget without mutating state."""
    messages = state.get("messages", [])
    if not messages:
        return END

    last = messages[-1]
    has_tool_calls = hasattr(last, "tool_calls") and last.tool_calls

    if has_tool_calls:
        budget = state.get("remaining_steps", 0)
        if budget > 0:
            logger.info(f"agent | routing to decrement_budget | remaining_steps={budget}")
            return "decrement_budget"
        logger.warning("agent | budget exhausted — forcing conversational response")
        return END

    logger.info("agent | routing to END (conversational response)")
    return END


def decrement_budget_node(state: InteractiveIngestionState) -> dict:
    """Decrements the tool call budget. Runs when route_agent_action finds tool_calls with budget > 0."""
    budget = state.get("remaining_steps", 0)
    logger.info(f"budget | decrementing: {budget} → {budget - 1}")
    return {"remaining_steps": budget - 1}


# ─── START Router ─────────────────────────────────────────────

def route_start(state: InteractiveIngestionState) -> str:
    """Route to triage on first run (profile_data empty), bypass to agent on subsequent turns."""
    if state.get("profile_data") is not None:
        logger.debug("start | profile_data exists — bypassing triage")
        return "agent"
    return "triage"


# ─── Graph Assembly ──────────────────────────────────────────────

builder = StateGraph(InteractiveIngestionState)

# Register tools with ToolNode
agent_tools = [profile_file, extract_categories, map_attributes, extract_references, build_products, render_templates, enrich_descriptions, merge_duplicates, merge_sheets_programmatically]
tool_node = ToolNode(agent_tools)

builder.add_node("triage", triage_interactive)
builder.add_node("agent", agent_reason_node)
builder.add_node("decrement_budget", decrement_budget_node)
builder.add_node("execute_tools", tool_node)

builder.add_conditional_edges(
    START,
    route_start,
    {"triage": "triage", "agent": "agent"},
)
builder.add_edge("triage", "agent")
builder.add_edge("decrement_budget", "execute_tools")
builder.add_edge("execute_tools", "agent")

builder.add_conditional_edges(
    "agent",
    route_agent_action,
    {"decrement_budget": "decrement_budget", END: END},
)

postgres_uri = os.getenv("POSTGRES_URI")
if postgres_uri:
    from langgraph.checkpoint.postgres import PostgresSaver
    from psycopg_pool import ConnectionPool
    pool = ConnectionPool(conninfo=postgres_uri, max_size=5)
    checkpointer = PostgresSaver(pool)
    checkpointer.setup()
else:
    checkpointer = MemorySaver()

interactive_graph = builder.compile(
    checkpointer=checkpointer,
)
