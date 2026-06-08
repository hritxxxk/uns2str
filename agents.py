import os
import json
import openpyxl
from dotenv import load_dotenv
from google import genai
from helpers import *
from state import ColumnMapping
from tools.profiling import detect_data_sheet, profile_columns, detect_category_structure
from tools.mapping import normalize_mapping, validate_mapping, build_attribute_definitions
from tools.references import extract_reference_values
from tools.rendering import render_category_xlsx, render_attribute_xlsx, render_reference_xlsx, render_product_xlsx

load_dotenv()
client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))

PIM_DEFAULTS = ["sku_name", "code"]

MAPPING_PROMPT_TEMPLATE = """You are a PIM data mapping expert. Map each source column to a PIM attribute.

Return a JSON object with key "mappings" containing an array of objects.
Each object has these fields:
  source_column (str): original column name
  target_attribute (str): snake_case PIM attribute name
  attribute_type (str): Textbox | Dropdown | RichText | Textarea | MultiSelect | Date | Time
  attribute_data_type (str): varchar | int | float | boolean | date
  constraint (bool): true only if dropdown or multiselect
  length (int): max characters
  mandatory (bool): true for identity or legal fields
  attribute_group (str): e.g. "Product Identification", "Pricing", "Classification", "Media", "Technical Specs", "Brand & Origin"
  confidence (float): 0.0 to 1.0

RULES:

1. target_attribute must be snake_case — e.g. "product_name", "item_code", "mrp", "colour", "brand", "gender", "size", "image_url". Do NOT copy source column names as-is.

2. Use column semantics + stats to decide attribute_type:
   - SKU, code, id fields → Textbox, varchar, mandatory=true
   - Brand, colour, size, gender, season, type, category → Dropdown, constraint=true
   - Product name / title → Textbox, mandatory=true
   - Description → RichText, varchar, length=65536
   - Price, MRP, cost → Textbox, float
   - Image/photo/img URLs → Textbox, varchar, length=2048
   - Tags, features, materials → MultiSelect, constraint=true
   - Date fields → Date, date

3. The PIM already has these default attributes: sku_name, code, description, mrp, brand. Do NOT recreate them — map source columns TO them instead (e.g. "Product Name" → "sku_name", not "product_name").

4. constraint=true ONLY for attributes with predefined selectable values.

5. mandatory=true ONLY for: sku, code, product_name, mrp.

Source columns with stats:
{profile_text}

Column metadata notes (data types, constraints, defaults above the header):
{metadata_text}

Sample rows:
{sample_text}"""

MAX_PROFILE_COLS = 150


def build_mapping_prompt(profiles, sample_rows, metadata=None):
    capped = profiles[:MAX_PROFILE_COLS]
    trimmed = [{k: v for k, v in p.items() if k != "unique_values"} for p in capped]
    for p in trimmed:
        if len(p.get("sample", [])) > 2:
            p["sample"] = p["sample"][:2]
    return MAPPING_PROMPT_TEMPLATE.format(
        profile_text=json.dumps(trimmed, indent=2),
        sample_text=json.dumps(sample_rows[:3], indent=2),
        metadata_text=json.dumps(metadata, indent=2) if metadata else "None"
    )


def call_llm(prompt):
    response = client.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=prompt,
        config={"response_mime_type": "application/json"}
    )
    return json.loads(response.text)


def _safe_json_parse(text):
    """Parse LLM JSON response, handling extra text gracefully."""
    import re as _re
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    decoder = json.JSONDecoder()
    for pattern in (r'\{.*?\}', r'\[.*?\]'):
        for match in _re.finditer(pattern, text, _re.DOTALL):
            try:
                return decoder.decode(match.group())
            except json.JSONDecodeError:
                continue
    return {}


def parse_mapping_response(raw):
    if isinstance(raw, list):
        return raw
    if isinstance(raw, dict):
        for key in ("mappings", "attributes", "columns"):
            if key in raw and isinstance(raw[key], list):
                return raw[key]
    raise ValueError(f"Cannot extract mappings from: {type(raw)}")


def cache_mapping(fingerprint, mappings):
    save_cached_mapping(fingerprint, [m.model_dump() for m in mappings])


def avg_confidence(mappings):
    if not mappings:
        return 0.0
    return sum(m.confidence for m in mappings) / len(mappings)


# ─── Graph nodes (thin orchestrators) ────────────────────────────

def fingerprint_source(state):
    
    rows = read_file(state["source_path"], state.get("sheet_name"))
    headers, _ = get_headers_and_data(rows)
    fp = fingerprint_headers(headers)
    cached = load_cached_mapping(fp)
    state["fingerprint"] = fp
    state["headers"] = headers
    state["is_known_schema"] = cached is not None
    if cached:
        state["mapping"] = [ColumnMapping(**m) for m in cached]
    else:
        state["mapping"] = []
    return state


def detect_header_via_llm(rows):
    preview = json.dumps([{f"col_{j}": str(c)[:40] for j, c in enumerate(row[:20]) if c is not None and str(c).strip()} for row in rows[:15]], indent=2)
    prompt = f"""Given the first 15 rows of a spreadsheet, identify:
1. Which row index (0-based) contains the column headers
2. Which row index (0-based) does the actual product data start at (skipping header and any metadata/description/constraint rows between the header and data).

Rows:
{preview}
Return JSON: {{"header_row": int, "data_start_row": int}}"""
    response = client.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=prompt,
        config={"response_mime_type": "application/json"}
    )
    result = json.loads(response.text)
    if isinstance(result, list):
        result = result[0] if result else {}
    return result.get("header_row", 0), result.get("data_start_row", len(rows))


def profile_source(state):
    if not state.get("sheet_name"):
        result = detect_data_sheet.invoke({"path": state["source_path"]})
        state["sheet_name"] = result["sheet"]
        print(f"  Auto-detected sheet: '{result['sheet']}' ({result['cells']} cells)")

    rows = read_file(state["source_path"], state.get("sheet_name"))
    header_row, data_start_row = detect_header_via_llm(rows)
    headers, data = get_headers_and_data(rows, header_row)
    if data_start_row < header_row + 1:
        data_start_row = header_row + 1
    data = rows[data_start_row:]

    state["headers"] = headers
    state["header_row"] = header_row
    state["data_start_row"] = data_start_row
    state["metadata"] = [{headers[j]: str(rows[mr][j])[:60] for j in range(min(len(headers), len(rows[mr]))) if rows[mr][j] is not None and str(rows[mr][j]).strip()} for mr in range(header_row)]
    state["profiles"] = profile_columns.invoke({"headers": headers, "rows": data})
    state["row_count"] = len(data)
    state["sample_rows"] = [dict(zip(headers, row)) for row in data[:5]]
    state["category_candidates"] = detect_category_structure.invoke({"path": state["source_path"], "data_sheet": state["sheet_name"]})
    state["category_path_config"] = {}
    state["category_hierarchy"] = []
    return state


def map_columns(state):
    if state["is_known_schema"]:
        return state

    prompt = build_mapping_prompt(state["profiles"], state.get("sample_rows", []), state.get("metadata"))
    raw = call_llm(prompt)
    extracted = parse_mapping_response(raw)
    normalized = normalize_mapping.invoke({"raw_list": extracted})
    parsed = validate_mapping.invoke({"raw_list": normalized})

    state["mapping"] = parsed
    state["mapping_requires_review"] = avg_confidence(parsed) < 0.75
    cache_mapping(state["fingerprint"], parsed)
    return state


def _validate_paths(paths):
    if not paths or len(paths) < 2:
        return False, "Less than 2 paths found"
    prompt = f"""Validate these category paths. They are VALID if they form a parent→child hierarchy.

ACCEPT paths that have:
- Prefix codes like "P_Product", "N_Item", "C_Code" (these are labels, not IDs)
- Mixed naming conventions

REJECT paths that have:
- True duplicate levels like "A > A > A" where same name repeats
- Garbage data like "t1 > temp > temp"
- Single values that aren't hierarchical

Sample paths:
{json.dumps(list(paths)[:10], indent=2)}

Return JSON: {{"is_valid": bool, "reason": "short explanation"}}"""
    resp = client.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=prompt,
        config={"response_mime_type": "application/json"}
    )
    r = json.loads(resp.text)
    if isinstance(r, list):
        r = r[0] if r else {}
    return r.get("is_valid", False), r.get("reason", "")


def _build_paths_from_columns(headers, rows, col_indices, sep=" > "):
    paths = set()
    for row in rows:
        parts = [str(row[i]).strip() for i in col_indices if i < len(row) and row[i] is not None and str(row[i]).strip()]
        if len(parts) >= 2:
            paths.add(sep.join(parts))
    return paths


def _strategy_hierarchy_sheet(state):
    candidates = state.get("category_candidates", [])
    if not candidates:
        return None
    trimmed = [{k: v for k, v in c.items()} for c in candidates[:3]]
    for t in trimmed:
        if len(t.get("headers", [])) > 20:
            t["headers"] = t["headers"][:20]
        if len(t.get("rows", [])) > 2:
            t["rows"] = t["rows"][:2]
    prompt = f"""Given sheets with potential hierarchy data, decide which columns form a category path.
Return JSON: {{"sheet": str, "columns": [str], "skip_code_columns": bool}}
Rules: Pick columns forming parent→child chain. Skip code/id/num columns.
Candidates:
{json.dumps(trimmed, indent=2)}"""
    resp = client.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=prompt,
        config={"response_mime_type": "application/json"}
    )
    r = _safe_json_parse(resp.text)
    if isinstance(r, list):
        r = r[0] if r else {}
    sheet_name = r.get("sheet", "")
    columns = r.get("columns", [])
    skip_codes = r.get("skip_code_columns", True)
    candidate = next((c for c in candidates if c["sheet"] == sheet_name), None)
    if not candidate or not columns:
        return None
    all_h = candidate["headers"]
    indices = [all_h.index(c) for c in columns if c in all_h]
    if len(indices) < 2:
        return None
    wb = openpyxl.load_workbook(state["source_path"], read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return _build_paths_from_columns(all_h, rows, indices)


def _strategy_level_columns(state):
    headers = state.get("headers", [])
    sample = state.get("sample_rows", [{}])[0] if state.get("sample_rows") else {}
    from tools.profiling import profile_columns
    rows = list(read_file(state["source_path"], state.get("sheet_name")))
    headers2, data2 = get_headers_and_data(rows, state.get("header_row", 0))
    dr = state.get("data_start_row", state.get("header_row", 0) + 1)
    if dr < state.get("header_row", 0) + 1:
        dr = state.get("header_row", 0) + 1
    cols = profile_columns.invoke({"headers": headers2, "rows": rows[dr:]})
    col_info = {c["name"]: {"unique": c["unique"], "sample": c["sample"][:2]} for c in cols if c["non_null"] > 0}
    prompt = f"""Pick columns forming a product category hierarchy (broad → narrow groupings).
Exclude: identifiers (SKU, codes, IDs), prices, descriptions, images, dates, names, colors, sizes, statuses, and columns with identical unique value sets.
Return JSON: {{"columns": ["col1", "col2", ...]}}
Columns with unique counts and samples:
{json.dumps(col_info, indent=2)}"""
    resp = client.models.generate_content(model="gemini-2.5-flash-lite", contents=prompt, config={"response_mime_type": "application/json"})
    r = _safe_json_parse(resp.text)
    if isinstance(r, list):
        r = r[0] if r else {}
    chosen = r.get("columns", [])
    indices = [headers.index(c) for c in chosen if c in headers]
    if len(indices) < 2:
        return None
    rows = list(read_file(state["source_path"], state.get("sheet_name")))
    _, data = get_headers_and_data(rows, state.get("header_row", 0))
    dr = state.get("data_start_row", state.get("header_row", 0) + 1)
    if dr < state.get("header_row", 0) + 1:
        dr = state.get("header_row", 0) + 1
    data = rows[dr:]
    return _build_paths_from_columns(headers, data, indices)


def _strategy_single_column(state):
    headers = state.get("headers", [])
    cat_col = next((i for i, h in enumerate(headers) if h.lower() in ("category", "categories", "product type")), None)
    if cat_col is None:
        return None
    rows = list(read_file(state["source_path"], state.get("sheet_name")))
    _, data = get_headers_and_data(rows, state.get("header_row", 0))
    dr = state.get("data_start_row", state.get("header_row", 0) + 1)
    if dr < state.get("header_row", 0) + 1:
        dr = state.get("header_row", 0) + 1
    data = rows[dr:]
    samples = list(set(str(row[cat_col]).strip() for row in data[:50] if row[cat_col] is not None and str(row[cat_col]).strip()))[:10]
    prompt = f"""Column '{headers[cat_col]}' contains category data. Sample values:
{json.dumps(samples, indent=2)}
Determine: multi_value_separator (if one cell has multiple categories), path_separator (between levels), split_cells (true/false).
Return JSON: {{"multi_value_separator": str or null, "path_separator": str, "split_cells": bool}}"""
    resp = client.models.generate_content(model="gemini-2.5-flash-lite", contents=prompt, config={"response_mime_type": "application/json"})
    r = _safe_json_parse(resp.text)
    if isinstance(r, list):
        r = r[0] if r else {}
    multi_sep = r.get("multi_value_separator")
    path_sep = r.get("path_separator", ">")
    split_cells = r.get("split_cells", False)
    paths = set()
    for row in data:
        val = str(row[cat_col]).strip() if cat_col < len(row) and row[cat_col] is not None else ""
        if not val:
            continue
        if split_cells and multi_sep and multi_sep in val:
            for chunk in val.split(multi_sep):
                chunk = chunk.strip()
                if not chunk:
                    continue
                if path_sep in chunk:
                    parts = [p.strip() for p in chunk.split(path_sep) if p.strip()]
                    if len(parts) >= 2:
                        paths.add(" > ".join(parts))
        elif path_sep in val:
            parts = [p.strip() for p in val.split(path_sep) if p.strip()]
            if len(parts) >= 2:
                paths.add(" > ".join(parts))
    return paths if len(paths) >= 2 else None


# ─── Declarative Recipe Strategy ──────────────────────────────

def _strategy_declarative_recipe(state):
    """Primary strategy: LLM writes a declarative recipe, Python executes it on 100% of rows.

    Handles 4 extraction modes:
    - level_columns: CATEGORY1, CATEGORY2 → path per row (handles empty middle levels)
    - synthesized: Department + Gender → "Activewear > Women"
    - lookup: Code "F-SH-01" → resolved from another sheet
    - single_column: "Category > Subcategory" → parsed by separator

    Self-healing pass merges near-duplicates ("Apperal" → "Apparel").
    """
    headers = state.get("headers", [])
    if not headers:
        return None

    # ── Profile columns with unique counts & samples ───────────
    rows = list(read_file(state["source_path"], state.get("sheet_name")))
    hr = state.get("header_row", 0)
    dr = state.get("data_start_row", hr + 1)
    if dr < hr + 1:
        dr = hr + 1
    data = rows[dr:]

    col_profiles = []
    for i, h in enumerate(headers):
        vals = [str(row[i]).strip() for row in data
                if i < len(row) and row[i] is not None and str(row[i]).strip()]
        if vals:
            unique = sorted(set(vals))
            col_profiles.append({
                "name": h,
                "unique_count": len(unique),
                "samples": unique[:5],
                "non_null": len(vals),
            })

    # ── Check if any other sheets exist for lookup ─────────────
    other_sheets = []
    ext = os.path.splitext(state["source_path"])[1].lower()
    if ext in (".xlsx", ".xls"):
        try:
            wb = openpyxl.load_workbook(state["source_path"], read_only=True, data_only=True)
            other_sheets = [s for s in wb.sheetnames if s != state.get("sheet_name")]
            wb.close()
        except Exception:
            pass

    # ── LLM writes the recipe ──────────────────────────────────
    prompt = f"""You are a category extraction expert. Analyze these column profiles and decide the best way to extract a clean product category hierarchy.

File: {os.path.basename(state['source_path'])}
Current sheet: {state.get('sheet_name', 'auto')}
Other sheets available: {other_sheets if other_sheets else 'none'}
Columns profiled ({len(col_profiles)} total):
{json.dumps(col_profiles[:40], indent=2)}

Edge cases to consider:
1. Missing middle levels (e.g. "Apparel > > t-shirts" — skip the empty level)
2. Categories spread across columns (e.g. Department + Gender → "Activewear > Women")
3. Category codes that need lookup (e.g. "F-SH-01" needs resolution from another sheet)
4. Single column with path separators (e.g. "Apparel/Mens/T-Shirts")
5. No obvious category columns — infer from attribute columns

Return a JSON recipe:
{{
  "strategy": "level_columns" | "synthesized" | "lookup" | "single_column" | "infer_from_attributes",
  "hierarchy_columns": ["col1", "col2", ...],
  "separator": " > ",
  "skip_empty_levels": true,
  "default_fill": "",
  "multi_value_separator": null,
  "path_separator": ">",
  "synthesize_with": null,
  "lookup_sheet": null,
  "lookup_code_column": null,
  "lookup_value_column": null,
  "lookup_path_separator": " > ",
  "conversational_explanation": "A short user-facing explanation of how I determined the hierarchy."
}}

Strategy rules:
- level_columns: Use when you see multiple columns like CATEGORY1-4 forming a hierarchy. hierarchy_columns = the level columns in order.
- synthesized: Use when no single category column exists but Department + Gender etc. can be joined. synthesize_with = joining word like " - ".
- lookup: Use when a column has short codes (F-SH-01). Set lookup_sheet + code/value columns.
- single_column: Use when one column contains paths like "Apparel > Mens > T-Shirts". Set path_separator and optionally multi_value_separator.
- infer_from_attributes: Last resort — pick columns that seem hierarchical. May use only 1 column.

IMPORTANT: skip_empty_levels=true handles the dirty intermediate case — empty/missing cells between valid levels are collapsed."""
    
    resp = client.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=prompt,
        config={"response_mime_type": "application/json"},
    )
    recipe = _safe_json_parse(resp.text)
    if isinstance(recipe, list):
        recipe = recipe[0] if recipe else {}

    strategy = recipe.get("strategy", "level_columns")
    hierarchy_cols = recipe.get("hierarchy_columns", [])
    sep = recipe.get("separator", " > ")
    skip_empty = recipe.get("skip_empty_levels", True)
    default_fill = recipe.get("default_fill", "")
    multi_sep = recipe.get("multi_value_separator")
    path_sep = recipe.get("path_separator", ">")

    if not hierarchy_cols:
        return None

    # ── Execute the recipe on 100% of rows ─────────────────────
    paths = set()

    if strategy == "level_columns":
        col_indices = [headers.index(c) for c in hierarchy_cols if c in headers]
        if len(col_indices) < 2:
            return None
        for row in data:
            parts = []
            for idx in col_indices:
                val = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ""
                if skip_empty and not val:
                    continue
                if val:
                    parts.append(val)
                elif not skip_empty and default_fill:
                    parts.append(default_fill)
            if len(parts) >= 2:
                paths.add(sep.join(parts))

    elif strategy == "synthesized":
        col_indices = [headers.index(c) for c in hierarchy_cols if c in headers]
        if len(col_indices) < 2:
            return None
        join_sep = recipe.get("synthesize_with", " - ") or " - "
        for row in data:
            parts = []
            for idx in col_indices:
                val = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ""
                if val:
                    parts.append(val)
            if len(parts) >= 2:
                paths.add(sep.join(parts))

    elif strategy == "lookup":
        lookup_sheet = recipe.get("lookup_sheet")
        code_col = recipe.get("lookup_code_column", hierarchy_cols[0]) if hierarchy_cols else None
        val_col = recipe.get("lookup_value_column", "category_path")
        lookup_sep = recipe.get("lookup_path_separator", " > ")

        # Build lookup map from the other sheet
        lookup_map = {}
        if lookup_sheet and other_sheets:
            try:
                wb = openpyxl.load_workbook(state["source_path"], read_only=True, data_only=True)
                if lookup_sheet in wb.sheetnames:
                    ws = wb[lookup_sheet]
                    lookup_headers = [str(c) if c else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
                    code_idx = 0  # default first col
                    val_idx = 1   # default second col
                    for ci, lh in enumerate(lookup_headers):
                        lh_lower = lh.lower().strip()
                        if "code" in lh_lower or "id" in lh_lower:
                            code_idx = ci
                        if "path" in lh_lower or "category" in lh_lower or "value" in lh_lower or "name" in lh_lower:
                            val_idx = ci
                    for lookup_row in ws.iter_rows(min_row=2, values_only=True):
                        code = str(lookup_row[code_idx]).strip() if code_idx < len(lookup_row) and lookup_row[code_idx] else ""
                        val = str(lookup_row[val_idx]).strip() if val_idx < len(lookup_row) and lookup_row[val_idx] else ""
                        if code and val:
                            lookup_map[code] = val
                wb.close()
            except Exception:
                pass

        col_idx = headers.index(code_col) if code_col in headers else 0
        for row in data:
            code = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ""
            if code in lookup_map:
                paths.add(lookup_map[code])
            elif code:
                # If code not in lookup, keep raw as fallback
                paths.add(code)

    elif strategy == "single_column":
        cat_col = next((i for i, h in enumerate(headers) if h in hierarchy_cols or h.lower() in ("category", "categories", "product type")), None)
        if cat_col is None:
            return None
        split_cells = recipe.get("split_cells", False)
        for row in data:
            val = str(row[cat_col]).strip() if cat_col < len(row) and row[cat_col] is not None else ""
            if not val:
                continue
            if split_cells and multi_sep and multi_sep in val:
                for chunk in val.split(multi_sep):
                    chunk = chunk.strip()
                    if not chunk:
                        continue
                    if path_sep in chunk:
                        parts = [p.strip() for p in chunk.split(path_sep) if p.strip()]
                        if len(parts) >= 2:
                            paths.add(sep.join(parts))
            elif path_sep in val:
                parts = [p.strip() for p in val.split(path_sep) if p.strip()]
                if len(parts) >= 2:
                    paths.add(sep.join(parts))

    elif strategy == "infer_from_attributes":
        col_indices = [headers.index(c) for c in hierarchy_cols if c in headers]
        if len(col_indices) < 2:
            return None
        for row in data:
            parts = [str(row[idx]).strip() for idx in col_indices
                     if idx < len(row) and row[idx] is not None and str(row[idx]).strip()]
            if len(parts) >= 2:
                paths.add(sep.join(parts))

    if len(paths) < 2:
        return None

    # ── Self-healing: normalize near-duplicates ────────────────
    healed = _heal_category_paths(paths)
    state["category_reasoning"] = recipe.get("conversational_explanation", "Discovered from column analysis.")
    return healed


# ─── Self-Healing ───────────────────────────────────────────────

def _heal_category_paths(paths: set, threshold: float = 0.85) -> list:
    """Fuzzy-merge near-duplicate category paths.

    Uses simple token overlap + case normalization.
    "Apperal > Mens > T-Shirts" and "Apparel > Men > T-Shirts"
    → "Apparel > Men > T-Shirts" (most common spelling wins).
    """
    import re as _re

    def tokenize(p: str) -> list[str]:
        return _re.sub(r'\s+', ' ', p.strip().lower()).split(" > ")

    def overlap(a: list[str], b: list[str]) -> float:
        if not a or not b or len(a) != len(b):
            return 0.0
        matches = sum(1 for i in range(len(a)) if a[i] == b[i] or
                      (len(a[i]) > 2 and len(b[i]) > 2 and
                       (a[i].startswith(b[i]) or b[i].startswith(a[i]) or
                        (len(set(a[i]) & set(b[i])) / max(len(set(a[i]) | set(b[i])), 1) > 0.7))))
        return matches / len(a)

    sorted_paths = sorted(paths)
    merged = []
    used = set()

    for i, p in enumerate(sorted_paths):
        if i in used:
            continue
        group = [p]
        used.add(i)
        tok_i = tokenize(p)
        for j in range(i + 1, len(sorted_paths)):
            if j in used:
                continue
            tok_j = tokenize(sorted_paths[j])
            if len(tok_i) == len(tok_j) and overlap(tok_i, tok_j) >= threshold:
                group.append(sorted_paths[j])
                used.add(j)

        # Pick the longest/most common form as canonical
        canonical = max(set(group), key=lambda x: (len(x), group.count(x)))
        merged.append(canonical)

    return sorted(merged)


def _strategy_infer_from_attributes(state):
    headers = state.get("headers", [])
    sample = state.get("sample_rows", [{}])[0] if state.get("sample_rows") else {}
    preview = {h: str(sample.get(h, ""))[:30] for h in headers if h.strip() and str(sample.get(h, "")).strip()}
    prompt = f"""Pick columns forming a product category hierarchy (broad → narrow groupings).
Exclude: identifiers (SKU, codes, IDs), prices, descriptions, images, dates, names, colors, sizes, statuses.
Return JSON: {{"columns": ["col1", "col2", ...]}}
Columns:
{json.dumps(preview, indent=2)}"""
    resp = client.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=prompt,
        config={"response_mime_type": "application/json"}
    )
    try:
        r = json.loads(resp.text)
    except json.JSONDecodeError:
        import re as _re
        match = _re.search(r'\{.*\}', resp.text, _re.DOTALL)
        r = json.loads(match.group()) if match else {}
    if isinstance(r, list):
        r = r[0] if r else {}
    chosen = r.get("columns", [])
    indices = [headers.index(c) for c in chosen if c in headers]
    if len(indices) < 2:
        return None
    rows = list(read_file(state["source_path"], state.get("sheet_name")))
    _, data = get_headers_and_data(rows, state.get("header_row", 0))
    dr = state.get("data_start_row", state.get("header_row", 0) + 1)
    if dr < state.get("header_row", 0) + 1:
        dr = state.get("header_row", 0) + 1
    data = rows[dr:]
    return _build_paths_from_columns(headers, data, indices)


def resolve_category_paths(state):
    strategies = [
        ("Declarative recipe (AI-generated)", _strategy_declarative_recipe),
        ("Hierarchy sheet", _strategy_hierarchy_sheet),
        ("Level columns (CATEGORY1-4)", _strategy_level_columns),
        ("Single category column", _strategy_single_column),
        ("Inferred from product attributes", _strategy_infer_from_attributes),
    ]

    for name, strategy in strategies:
        result = strategy(state)
        if result is None:
            continue
        valid, reason = _validate_paths(result)
        print(f"  Category strategy '{name}': {'✅' if valid else '❌'} {reason}")
        if valid:
            state["category_hierarchy"] = sorted(result)
            return state

    state["need_user_input"] = True
    print("  ⚠ Could not determine category paths. Set need_user_input=True")
    return state


def build_attributes(state):
    defs = build_attribute_definitions.invoke({"mappings": state["mapping"]})
    state["attribute_definitions"] = defs
    return state


def collect_references(state):
    raw_mappings = [{"source_column": m.source_column, "target_attribute": m.target_attribute, "attribute_type": m.attribute_type} for m in state["mapping"]]
    refs = extract_reference_values.invoke({"mappings": raw_mappings, "profiles": state["profiles"]})
    state["reference_values"] = refs
    return state


def fill_templates(state):
    os.makedirs("output", exist_ok=True)
    fp = state["fingerprint"]
    files = {}

    if state.get("category_hierarchy"):
        wb = render_category_xlsx.invoke({"paths": state["category_hierarchy"]})
        wb.save(f"output/{fp}_category.xlsx")
        wb.close()
        files["category"] = f"output/{fp}_category.xlsx"

    wb = render_attribute_xlsx.invoke({"defs": state["attribute_definitions"]})
    wb.save(f"output/{fp}_attribute.xlsx")
    wb.close()
    files["attribute"] = f"output/{fp}_attribute.xlsx"

    if state.get("reference_values"):
        wb = render_reference_xlsx.invoke({"refs": state["reference_values"]})
        wb.save(f"output/{fp}_reference.xlsx")
        wb.close()
        files["reference"] = f"output/{fp}_reference.xlsx"

    rows = read_file(state["source_path"], state.get("sheet_name"))
    hr = state.get("header_row", 0)
    headers, data = get_headers_and_data(rows, hr)
    dr = state.get("data_start_row", hr + 1)
    if dr < hr + 1:
        dr = hr + 1
    data = rows[dr:]
    img_cols = extract_image_columns(headers)
    mapping_list = [{"source_column": m.source_column, "target_attribute": m.target_attribute} for m in state["mapping"]]
    attr_names = [m.get("target_attribute", m.get("source_column")) for m in mapping_list]

    product_rows = build_product_rows(headers, data, mapping_list, img_cols)
    wb = render_product_xlsx.invoke({"rows": product_rows, "attr_names": attr_names})
    wb.save(f"output/{fp}_product.xlsx")
    wb.close()
    files["product"] = f"output/{fp}_product.xlsx"

    state["output_files"] = files
    return state


# ─── VinGPT Nodes ─────────────────────────────────────────────
import json
import os

from state import PIM_DEFAULTS


ANALYZE_PROMPT = """You are a friendly data analyst helping a non-technical user prepare their spreadsheet for a PIM system.

Given the file info and sample data below, identify:
1. Which columns match the standard fields: {standard_fields}
2. Which columns are custom/dynamic attributes the user wants to keep
3. Any missing standard fields that have no matching column

Return JSON:
{{
  "core": [{{"column": "src name", "target": "sku", "confidence": 95}}],
  "custom_columns": ["Col1", "Col2"],
  "missing_core": []
}}
Confidence: 0-100 rating. Below 85 means uncertain and needs user confirmation.

File: {filename}
Columns found: {columns}
Sample data:
{samples}
"""


def analyze_and_ask(state: dict) -> dict:
    profile = state.get("profile_data")
    if not profile:
        state["pending_questions"] = ["No profile data found. Please provide a valid file."]
        return state

    headers = profile.get("headers", [])
    samples = profile.get("sample_rows", [])
    row_count = profile.get("row_count", 0)

    prompt = ANALYZE_PROMPT.format(
        standard_fields=", ".join(PIM_DEFAULTS),
        filename=os.path.basename(state.get("file_path", "")),
        columns=", ".join(headers),
        samples=json.dumps(samples[:3], indent=2) if samples else "No sample rows",
    )

    from google import genai as _genai
    client = _genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
    resp = client.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=prompt,
        config={"response_mime_type": "application/json"},
    )

    try:
        result = json.loads(resp.text)
    except json.JSONDecodeError:
        import re as _re
        m = _re.search(r'\{.*\}', resp.text, _re.DOTALL)
        result = json.loads(m.group()) if m else {}

    core_list = result.get("core", []) if isinstance(result.get("core"), list) else []
    custom_cols = result.get("custom_columns", [])
    missing = result.get("missing_core", [])

    state["core_mappings"] = {item["target"]: item["column"] for item in core_list if item.get("column")}
    state["mapping_confidence"] = {item["target"]: item["confidence"] for item in core_list if item.get("column")}
    state["custom_mappings"] = {col: col for col in custom_cols}

    questions = []
    for item in core_list:
        if item.get("column"):
            questions.append({
                "id": f"core_{item['target']}",
                "type": "core",
                "target": item["target"],
                "column": item["column"],
                "confidence": item.get("confidence", 100),
                "text": f"I found '{item['column']}' — looks like it could be the **{item['target']}** field. Is that right? (Yes/No)",
            })
    for pim_key in missing:
        questions.append({
            "id": f"missing_{pim_key}",
            "type": "missing",
            "target": pim_key,
            "column": "",
            "confidence": 0,
            "text": f"I couldn't find a column that looks like **{pim_key}**. Do you want to leave it blank for now? (Yes/No)",
        })
    if custom_cols:
        col_list = "**, **".join(custom_cols[:5])
        suffix = f" and {len(custom_cols) - 5} more" if len(custom_cols) > 5 else ""
        questions.append({
            "id": "custom_attrs",
            "type": "custom",
            "target": "",
            "column": "",
            "confidence": 100,
            "columns": custom_cols,
            "text": f"I also see **{col_list}**{suffix} — these look like custom attributes specific to your products. Should I keep them as-is? (Yes/No)",
        })

    state["pending_questions"] = questions

    msg = (
        f"Great, I've scanned **{os.path.basename(state.get('file_path', ''))}** "
        f"({row_count} rows, {len(headers)} columns). "
        f"I have {len(questions)} questions before I proceed."
    )
    state.setdefault("messages", []).append({"role": "assistant", "content": msg})

    return state


def check_confidence(state: dict) -> dict:
    conf = state.get("mapping_confidence", {})
    questions = state.setdefault("pending_questions", [])

    low_conf_items = [(tgt, col, score) for tgt, col in state.get("core_mappings", {}).items()
                      if (score := conf.get(tgt, 100)) < 85]

    for target, column, score in low_conf_items:
        questions.append({
            "id": f"conf_{target}",
            "type": "core",
            "target": target,
            "column": column,
            "confidence": score,
            "text": f"I'm not very sure about **{column}** → **{target}** (confidence: {score}%). Can you confirm this is correct? (Yes/No)",
        })

    if low_conf_items:
        msg = f"I have {len(low_conf_items)} more quick question{'s' if len(low_conf_items) > 1 else ''} about mappings I'm unsure about."
        state.setdefault("messages", []).append({"role": "assistant", "content": msg})

    return state


# ─── ZIP Pre-Processor: Union Recipe ──────────────────────────

UNION_RECIPE_PROMPT = """You are a PIM data consolidation expert analyzing marketplace export files.

You are given the headers and sample rows from multiple source files that need to be merged into a single unified product catalog.

Your task: create a ConsolidationRecipe JSON that maps each source file's columns to standardized PIM attributes.

PIM target attributes (map source columns TO these):
- code (product SKU/identifier)
- sku_name (product title/name)
- mrp (price)
- brand
- category_path
- description
- size
- color
- material
- gender
- season
- image_1 through image_9

Source files and their headers:
{profiles_json}

For each file, identify which columns map to which PIM target attributes.
Different platforms use different names for the same thing:
- Shopify "Variant SKU" → code
- Flipkart "Style Code" → code
- Amazon "seller-sku" → code
- Magento "sku" → code
- WooCommerce "SKU" → code
- Meesho "Product Code" → code

Available transformations: strip_and_uppercase, strip_and_lowercase, to_float, strip_html, null

Return JSON:
{{
  "target_mappings": {{
    "code": {{
      "sources": {{"Shopify.csv": "Variant SKU", "Flipkart.xlsx": "Style Code"}},
      "transformation": "strip_and_uppercase"
    }},
    "sku_name": {{
      "sources": {{"Shopify.csv": "Title"}},
      "transformation": "strip_html"
    }},
    "mrp": {{
      "sources": {{"Shopify.csv": "Variant Price"}},
      "transformation": "to_float"
    }}
  }},
  "unified_headers": ["code", "sku_name", "mrp", "brand", "description", "size", "color", "category_path", "image_1", "image_2", "image_3"],
  "summary": "Combined Shopify and Flipkart exports into unified product catalog."
}}

IMPORTANT:
- Every source file must have at least its code column mapped.
- For columns that don't map to any PIM target, exclude them.
- Keep unified_headers concise: only include attributes that have at least one source mapping.
"""


def build_union_recipe(file_profiles: dict) -> dict:
    """Analyze headers from multiple source files and generate a ConsolidationRecipe.

    Args:
        file_profiles: dict from profile_files() — {filename: {headers, samples, row_count, ext}}

    Returns:
        ConsolidationRecipe dict with target_mappings and unified_headers.
    """
    profiles_json = json.dumps(file_profiles, indent=2)
    prompt = UNION_RECIPE_PROMPT.format(profiles_json=profiles_json)
    result = call_llm(prompt)

    if isinstance(result, list):
        result = result[0] if result else {}

    if "target_mappings" not in result:
        logger.warning("build_union_recipe: LLM did not return target_mappings")
        return {"target_mappings": {}, "unified_headers": [], "summary": "Failed to generate recipe."}

    return result
