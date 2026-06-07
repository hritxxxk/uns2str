import csv
import hashlib
import json
import os
import re

import openpyxl
from dotenv import load_dotenv
from langgraph.graph import StateGraph, START, END
from langgraph.checkpoint.memory import MemorySaver
from state import AgentState, IngestionOutput, ColumnMapping, MappingLLMResponse, PIM_DEFAULTS
from helpers import read_file, take_rows, get_headers_and_data, build_product_rows, extract_image_columns, fingerprint_headers, load_cached_mapping, save_cached_mapping, download_blank_template
from learning import fetch_similar_examples
from tools.mapping import build_attribute_definitions
from tools.references import extract_reference_values
from tools.rendering import render_all_templates


load_dotenv()
api_key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")


# ─── Triage Node ────────────────────────────────────────────────

def triage_source(state: dict) -> dict:
    path = state["source_path"]
    ext = os.path.splitext(path)[1].lower()

    sheets = []
    sheet_count = 0
    best_sheet = state.get("sheet_name") or ""
    first_rows = []
    total_row_count = 0

    if ext == ".csv":
        sheet_count = 1
        best_sheet = best_sheet or ""
        gen = read_file(path)
        first_rows = take_rows(gen, 20)
        total_row_count = 1 + sum(1 for _ in gen)
    elif ext in (".xlsx", ".xls"):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            sheet_count = len(sheets)
            if best_sheet and best_sheet in sheets:
                ws = wb[best_sheet]
            else:
                best_sheet = sheets[0]
                best_size = 0
                for sn in sheets:
                    ws = wb[sn]
                    first = next(ws.iter_rows(max_row=1, values_only=True), [])
                    cols = sum(1 for c in first if c is not None)
                    row_est = ws.max_row or 0
                    size = cols * row_est
                    if size > best_size:
                        best_size = size
                        best_sheet = sn
                        ws = wb[best_sheet]
            total_row_count = ws.max_row or 0
            wb.close()
            gen = read_file(path, best_sheet)
            first_rows = take_rows(gen, 20)
        except Exception:
            import xlrd
            xl = xlrd.open_workbook(path)
            sheets = xl.sheet_names()
            sheet_count = len(sheets)
            if best_sheet and best_sheet in sheets:
                idx = sheets.index(best_sheet)
            else:
                idx = 0
                best_sheet = sheets[0]
            ws = xl.sheet_by_index(idx)
            total_row_count = ws.nrows
            gen = read_file(path, best_sheet)
            first_rows = take_rows(gen, 20)
    else:
        gen = read_file(path, best_sheet or None)
        first_rows = take_rows(gen, 20)
        sheet_count = 1
        total_row_count = 1 + sum(1 for _ in gen)

    header_row = 0
    for i, row in enumerate(first_rows):
        cleaned = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if cleaned:
            header_row = i
            break

    headers = [str(c) if c is not None else "" for c in first_rows[header_row]]
    data_start_row = header_row + 1
    row_count = total_row_count - data_start_row
    if row_count < 0:
        row_count = 0

    column_count = len(headers)
    fingerprint = fingerprint_headers(headers)
    cached = load_cached_mapping(fingerprint)

    basic_profiles = [
        {"name": h, "col_index": i}
        for i, h in enumerate(headers) if h.strip()
    ]

    state["sheet_name"] = best_sheet
    state["fingerprint"] = fingerprint
    state["is_known_schema"] = cached is not None
    state["headers"] = headers
    state["header_row"] = header_row
    state["data_start_row"] = data_start_row
    state["row_count"] = row_count
    state["profiles"] = basic_profiles
    state["sheet_count"] = sheet_count
    state["column_count"] = column_count
    state["sheets"] = sheets

    if cached:
        state["mapping"] = [ColumnMapping(**m) for m in cached]

    return state


# ─── Category Resolution Node ───────────────────────────────────

def resolve_categories(state: dict) -> dict:
    if state.get("is_known_schema") and state.get("category_hierarchy"):
        return state
    from agents import resolve_category_paths
    resolve_category_paths(state)
    if not state.get("category_hierarchy"):
        state["need_user_input"] = False
    return state


# ─── Specialist Mapping Node ────────────────────────────────────

def map_columns_specialist(state: dict) -> dict:

    is_retry = bool(state.get("validation_message"))
    if not is_retry and state.get("is_known_schema") and state.get("mapping"):
        return state

    rows = read_file(state["source_path"], state.get("sheet_name"))
    header_row = state.get("header_row", 0)
    data_start = state.get("data_start_row", header_row + 1)
    headers = state.get("headers", [])

    for _ in range(data_start):
        try:
            next(rows)
        except StopIteration:
            break

    sample_rows_data = take_rows(rows, 5)
    sample_rows = []
    for row in sample_rows_data:
        sample = {}
        for i, h in enumerate(headers):
            if i < len(row) and row[i] is not None and str(row[i]).strip():
                val = str(row[i]).strip()[:80]
                if val:
                    sample[h] = val
        sample_rows.append(sample)

    few_shots = []
    seen_targets = set()
    for h in headers[:10]:
        vals = []
        for row in sample_rows_data:
            idx = headers.index(h)
            if idx < len(row) and row[idx] is not None and str(row[idx]).strip():
                vals.append(str(row[idx]).strip()[:40])
        if not vals:
            continue
        matches = fetch_similar_examples(h, vals, k=2)
        for m in matches:
            tgt = m["target_attribute"]
            if tgt and tgt not in seen_targets:
                few_shots.append(m)
                seen_targets.add(tgt)
                if len(few_shots) >= 5:
                    break
        if len(few_shots) >= 5:
            break

    prompt = f"""You are a PIM data mapping specialist. Map each source column to a PIM attribute.

File: {os.path.basename(state['source_path'])} ({state.get('row_count', 0)} rows)

Headers and sample values:
{json.dumps(sample_rows, indent=2)}

PIM defaults (do NOT recreate — map TO them): sku_name, code, mrp

Historical corrections for similar columns:
{chr(10).join(f'- \"{fs["column_name"]}\" → {fs["target_attribute"]} ({fs["attribute_type"]}, {fs["attribute_data_type"]}, mandatory={str(fs["mandatory"]).lower()})' for fs in few_shots) if few_shots else '(no historical corrections available)'}

Rules:
- target_attribute: snake_case
- Dropdown for brand/colour/size/gender/season/category (constraint=true)
- RichText for descriptions (length=65536)
- Textbox for codes/names/numbers/prices/images
- MultiSelect for multi-value tags/features (constraint=true)
- Textarea for attributes with >200 words
- Date for date fields
- mandatory=true only for sku/code/product_name/mrp
- Image URL columns: type=Textbox, length=2048, data_type=varchar
- confidence: 1.0 for clear matches, 0.5 for guesses, 0.0 for unknown

Identify system-critical columns by name (fill these fields):
- core_sku_column → which source column holds the SKU?
- core_code_column → which source column holds the product code?
- core_mrp_column → which source column holds the price/mrp?

Return mapping (every source column mapped), core_sku_column,
core_code_column, core_mrp_column, and needs_human_input."""

    if is_retry:
        prompt += f"\n\n---\n{state['validation_message']}\n---"

    from google import genai
    hclient = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))

    resp = hclient.models.generate_content(
        model="gemini-2.5-flash-lite",
        contents=prompt,
        config={
            "response_mime_type": "application/json",
            "response_schema": MappingLLMResponse
        }
    )

    result = json.loads(resp.text)
    if isinstance(result, list):
        result = result[0] if result else {}

    raw_mappings = result.get("mapping", [])
    parsed = []
    for m in raw_mappings:
        if isinstance(m, dict):
            parsed.append(ColumnMapping(**m))
        else:
            parsed.append(m)

    state["mapping"] = parsed
    state["core_column_detection"] = {
        "sku": result.get("core_sku_column", ""),
        "code": result.get("core_code_column", ""),
        "mrp": result.get("core_mrp_column", ""),
        "category": result.get("core_category_column", ""),
    }
    state["need_user_input"] = result.get("needs_human_input", False)

    if parsed:
        avg_conf = sum(m.confidence for m in parsed) / len(parsed)
        state["mapping_requires_review"] = avg_conf < 0.75

    return state


# ─── Evaluation Node ────────────────────────────────────────────

def _check_type_compatibility(samples: list[str], declared_type: str) -> list[str]:

    if declared_type in ("varchar", "varchar[]"):
        return []

    non_matching = []
    for v in samples:
        v = v.strip()
        if not v:
            continue
        if declared_type == "int":
            try:
                int(v)
            except ValueError:
                non_matching.append(v)
        elif declared_type == "float":
            cleaned = v.replace(",", "").replace("$", "").replace("€", "").replace("£", "").replace("₹", "")
            try:
                float(cleaned)
            except ValueError:
                non_matching.append(v)
        elif declared_type == "boolean":
            if v.lower() not in ("true", "false", "yes", "no", "0", "1", "y", "n", "t", "f"):
                non_matching.append(v)
        elif declared_type == "date":
            date_patterns = [
                r"^\d{4}-\d{2}-\d{2}$",
                r"^\d{2}/\d{2}/\d{4}$",
                r"^\d{2}-\d{2}-\d{4}$",
                r"^\d{4}/\d{2}/\d{2}$",
                r"^\d{2}\.\d{2}\.\d{4}$",
            ]
            if not any(re.match(p, v) for p in date_patterns):
                non_matching.append(v)
    return non_matching


def evaluate_mappings(state: dict) -> dict:

    if state.get("human_approved"):
        state["validation_errors"] = []
        state["validation_message"] = ""
        return state
    errors = []
    rows = read_file(state["source_path"], state.get("sheet_name"))
    header_row = state.get("header_row", 0)
    data_start = state.get("data_start_row", header_row + 1)
    headers = state.get("headers", [])

    for _ in range(data_start):
        try:
            next(rows)
        except StopIteration:
            break
    sample_rows_data = take_rows(rows, 10)

    mapping = state.get("mapping", [])
    core_cols = state.get("core_column_detection", {})

    # ── A. Type compatibility checks ─────────────────────────────
    for m in mapping:
        src = m.source_column
        declared_type = m.attribute_data_type
        target = m.target_attribute

        if src not in headers:
            errors.append({
                "field": target,
                "issue": f"source_column '{src}' not found in headers",
                "samples": []
            })
            continue

        col_idx = headers.index(src)
        samples = []
        for row in sample_rows_data:
            if col_idx < len(row) and row[col_idx] is not None and str(row[col_idx]).strip():
                samples.append(str(row[col_idx]).strip())
        if not samples:
            continue

        bad = _check_type_compatibility(samples, declared_type)
        if bad and len(bad) / max(len(samples), 1) >= 0.2:
            errors.append({
                "field": target,
                "issue": f"Type mismatch: declared '{declared_type}' but ≥20% of samples don't conform",
                "samples": bad[:5]
            })

    # ── B. Mandatory attribute checks ────────────────────────────
    mapped_targets = {m.target_attribute for m in mapping}
    for default in PIM_DEFAULTS:
        if default not in mapped_targets:
            errors.append({
                "field": default,
                "issue": f"Missing mandatory PIM default: '{default}' has no mapping",
                "samples": []
            })

    for key in ("sku", "code", "mrp"):
        if not core_cols.get(key):
            errors.append({
                "field": key,
                "issue": f"core_column_detection['{key}'] is empty — system-critical column not identified",
                "samples": []
            })

    state["validation_errors"] = errors

    if errors:
        cycle = state.get("correction_cycle", 0) + 1
        state["correction_cycle"] = cycle

        lines = [f"Validation failed (attempt {cycle}/3). Fix these issues:"]
        for err in errors:
            line = f"\n- {err['field']}: {err['issue']}"
            if err.get("samples"):
                line += f"\n  Offending samples: {err['samples']}"
            lines.append(line)
        lines.append("\nReturn a corrected IngestionOutput with all issues resolved.")
        state["validation_message"] = "\n".join(lines)
    else:
        state["validation_message"] = ""

    return state


def route_after_evaluation(state: dict) -> str:

    errors = state.get("validation_errors", [])
    cycle = state.get("correction_cycle", 0)

    if errors:
        return "retry" if cycle < 3 else "fail"

    if state.get("need_user_input"):
        return "halt"

    return "render"


# ─── Render Node ────────────────────────────────────────────────

def render_agent(state: dict) -> dict:

    fp = state.get("fingerprint") or fingerprint_headers(state.get("headers", []))
    mapping = state.get("mapping", [])
    headers = state.get("headers", [])
    cats = state.get("category_hierarchy", [])

    attr_defs = build_attribute_definitions.invoke({"mappings": mapping})

    mapping_dicts = [
        {"source_column": m.source_column, "target_attribute": m.target_attribute, "attribute_type": m.attribute_type}
        for m in mapping
    ]
    refs = extract_reference_values.invoke({"mappings": mapping_dicts, "profiles": state.get("profiles", [])})

    rows = read_file(state["source_path"], state.get("sheet_name"))
    hr = state.get("header_row", 0)
    dr = max(state.get("data_start_row", hr + 1), hr + 1)
    for _ in range(dr):
        try:
            next(rows)
        except StopIteration:
            break
    data = rows
    row_mappings = [{"source_column": m.source_column, "target_attribute": m.target_attribute} for m in mapping]
    img_cols = extract_image_columns(headers, row_mappings)
    attr_names = [m.get("target_attribute", m.get("source_column")) for m in row_mappings]
    product_rows = build_product_rows(headers, data, row_mappings, img_cols, state.get("core_column_detection"))

    files = render_all_templates.invoke({
        "fingerprint": fp,
        "category_hierarchy": cats,
        "attribute_definitions": attr_defs,
        "reference_values": refs,
        "headers": headers,
        "product_rows": product_rows,
        "attr_names": attr_names
    })

    state["attribute_definitions"] = attr_defs
    state["reference_values"] = refs
    state["output_files"] = files
    state["structured_response"] = IngestionOutput(
        status="success",
        fingerprint=fp,
        attribute_count=len(attr_defs),
        reference_count=len(refs),
        category_count=len(cats),
        output_files=list(files.values()),
        message=f"Generated {len(files)} output files for {os.path.basename(state.get('source_path', ''))}",
        mapping=[m.model_dump() for m in mapping],
        header_row=state.get("header_row", 0),
        data_start_row=state.get("data_start_row", 1),
        category_hierarchy=cats,
        sheet_name=state.get("sheet_name", ""),
        core_column_detection=state.get("core_column_detection", {}),
        needs_human_input=state.get("need_user_input", False),
    )

    return state


builder = StateGraph(AgentState)

builder.add_node("triage", triage_source)
builder.add_node("categories", resolve_categories)
builder.add_node("mapper", map_columns_specialist)
builder.add_node("evaluate", evaluate_mappings)
builder.add_node("render", render_agent)

builder.set_entry_point("triage")
builder.add_edge("triage", "categories")
builder.add_edge("categories", "mapper")
builder.add_edge("mapper", "evaluate")

builder.add_conditional_edges(
    "evaluate",
    route_after_evaluation,
    {"retry": "mapper", "fail": END, "halt": END, "render": "render"},
)

builder.add_edge("render", END)

postgres_uri = os.getenv("POSTGRES_URI")
if postgres_uri:
    from langgraph.checkpoint.postgres import PostgresSaver
    from psycopg_pool import ConnectionPool

    pool = ConnectionPool(conninfo=postgres_uri, max_size=5)
    checkpointer = PostgresSaver(pool)
    checkpointer.setup()
else:
    checkpointer = MemorySaver()

checkpointer = checkpointer.with_allowlist([
    ("state", "ColumnMapping"),
    ("state", "IngestionOutput"),
])

graph = builder.compile(
    checkpointer=checkpointer,
    interrupt_after=["evaluate"],
)

# ─── VinGPT Graph ───────────────────────────────────────────────

from state import IngestionState
from agents import analyze_and_ask, check_confidence

vingpt_builder = StateGraph(IngestionState)

def _route_questions(state: dict) -> str:
    if state.get("pending_questions"):
        return "human_input"
    return "render"

def _human_input_node(state: dict) -> dict:
    return state

def _render_vingpt(state: dict) -> dict:
    from tools.mapping import build_attribute_definitions
    from tools.references import extract_reference_values
    from tools.rendering import render_all_templates

    mapping_objs = []
    for target, col in state.get("core_mappings", {}).items():
        if col:
            mapping_objs.append(ColumnMapping(source_column=col, target_attribute=target, confidence=1.0))
    for col, preserved in state.get("custom_mappings", {}).items():
        mapping_objs.append(ColumnMapping(source_column=col, target_attribute=preserved, confidence=1.0))

    fp = fingerprint_headers(state.get("profile_data", {}).get("headers", []))
    headers = state.get("profile_data", {}).get("headers", [])
    cats = state.get("profile_data", {}).get("category_hierarchy", [])

    attr_defs = build_attribute_definitions.invoke({"mappings": mapping_objs})
    refs = extract_reference_values.invoke({
        "mappings": [{"source_column": m.source_column, "target_attribute": m.target_attribute, "attribute_type": m.attribute_type} for m in mapping_objs],
        "profiles": state.get("profile_data", {}).get("profiles", []),
    })
    rows = read_file(state["file_path"], state.get("sheet_name"))
    hr = state.get("profile_data", {}).get("header_row", 0)
    dr = hr + 1
    for _ in range(dr):
        try:
            next(rows)
        except StopIteration:
            break
    data = rows
    img_cols = extract_image_columns(headers)
    row_mappings = [{"source_column": m.source_column, "target_attribute": m.target_attribute} for m in mapping_objs]
    attr_names = [m.target_attribute for m in mapping_objs]
    product_rows = build_product_rows(headers, data, row_mappings, img_cols, state.get("core_mappings"))

    jwt = state.get("jwt_token", "")
    blank_cat = download_blank_template(jwt, "category") if jwt else ""
    blank_attr = download_blank_template(jwt, "attribute") if jwt else ""
    files = render_all_templates.invoke({
        "fingerprint": fp,
        "category_hierarchy": cats,
        "attribute_definitions": attr_defs,
        "reference_values": refs,
        "headers": headers,
        "product_rows": product_rows,
        "attr_names": attr_names,
        "blank_category_path": blank_cat,
        "blank_attribute_path": blank_attr,
    })

    state["generated_files"] = list(files.values())
    msg = f"All done! Generated {len(files)} PIM template files."
    state.setdefault("messages", []).append({"role": "assistant", "content": msg})
    return state

vingpt_builder.add_node("analyze", analyze_and_ask)
vingpt_builder.add_node("check_conf", check_confidence)
vingpt_builder.add_node("human_input", _human_input_node)
vingpt_builder.add_node("render", _render_vingpt)

vingpt_builder.add_edge(START, "analyze")
vingpt_builder.add_edge("analyze", "check_conf")

vingpt_builder.add_conditional_edges(
    "check_conf",
    _route_questions,
    {"human_input": "human_input", "render": "render"},
)

vingpt_builder.add_edge("human_input", "check_conf")
vingpt_builder.add_edge("render", END)

vingpt_graph = vingpt_builder.compile(
    checkpointer=checkpointer,
    interrupt_after=["human_input"],
)
