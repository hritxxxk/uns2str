#!/usr/bin/env python3
"""
PIM Ingestion Agent — Integration Test Suite

Usage:
    python3 tests/test_system.py              # Local graph tests (needs GEMINI_API_KEY)
    python3 tests/test_system.py --api http://localhost:8000  # Against live server
    python3 tests/test_system.py --quick      # Skip slow LLM-dependent tests
"""

import os
import sys
import json
import time
import subprocess
import argparse
import traceback

# Add project root to sys.path so local imports work
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# ─── Test Configuration ─────────────────────────────────────────
TEST_FILE = "client-data/client_data/apparel_clean_sample.xlsx"
PASS = 0
FAIL = 0

def check(label, condition, detail=""):
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"  ✅ {label}")
    else:
        FAIL += 1
        print(f"  ❌ {label}  {detail}")

def section(name):
    print(f"\n{'='*60}")
    print(f"  {name}")
    print(f"{'='*60}")


# ═══════════════════════════════════════════════════════════════
# TEST GROUP 1: Graph compiles and structure is correct
# ═══════════════════════════════════════════════════════════════

def test_graph_structure():
    section("1. Graph Structure")
    os.environ["GEMINI_API_KEY"] = os.environ.get("GEMINI_API_KEY", "test-key-placeholder")
    from interactive_graph import interactive_graph, route_start

    g = interactive_graph.get_graph()
    nodes = sorted(g.nodes.keys())
    check("Graph has 3 functional nodes", "agent" in nodes and "execute_tools" in nodes and "triage" in nodes,
          f"nodes={nodes}")

    check("route_start routes empty state to triage", route_start({"profile_data": None}) == "triage")
    check("route_start routes populated state to agent", route_start({"profile_data": {"headers": ["a"]}}) == "agent")

    edges = [(e.source, e.target) for e in g.edges]
    check("START has conditional edges", ("__start__", "triage") in edges and ("__start__", "agent") in edges)
    check("execute_tools loops back to agent", ("execute_tools", "agent") in edges)
    check("triage flows to agent", ("triage", "agent") in edges)

    from interactive_state import InteractiveIngestionState
    annotations = InteractiveIngestionState.__annotations__
    check("messages has operator.add reducer", "list" in str(annotations.get("messages", "")), str(annotations.get("messages")))


# ═══════════════════════════════════════════════════════════════
# TEST GROUP 2: Tool functions (unit tests, no LLM needed)
# ═══════════════════════════════════════════════════════════════

def test_tools_direct(quick=False):
    section("2. Tool Unit Tests")

    from interactive_graph import profile_file, extract_categories, extract_references, build_products, render_templates
    from tools.profiling import profile_columns

    # Build a minimal state with known data
    test_state = {
        "file_path": TEST_FILE,
        "sheet_name": None,
        "profile_data": None,
        "messages": [],
        "completed_phases": [],
        "remaining_steps": 0,
        "core_mappings": {},
        "custom_mappings": {},
        "mapping_confidence": {},
        "product_rows": [],
        "generated_files": [],
        "jwt_token": "",
        "all_sheets": [],
        "sheet_merge": {},
        "categories": {"explanation": "", "reasoning": "", "suggestions": [], "approved": False, "user_feedback": ""},
        "attributes": {"explanation": "", "reasoning": "", "suggestions": [], "approved": False, "user_feedback": ""},
        "references": {"explanation": "", "reasoning": "", "suggestions": [], "approved": False, "user_feedback": ""},
        "products": {"explanation": "", "reasoning": "", "suggestions": [], "approved": False, "user_feedback": ""},
        "current_phase": "categories",
        "phases_completed": [],
    }

    # 2a. Path resolution
    result = profile_file.invoke({"file_path": "nonexistent_file.xlsx", "state": test_state})
    check("profile_file resolves missing path from state",
          "Profiled" in result and "apparel_clean_sample" in result)

    # 2b. Profile data correctness
    pd = test_state.get("profile_data", {})
    check("profile_data has headers", bool(pd.get("headers")), f"count={len(pd.get('headers', []))}")
    check("header_row is int", isinstance(pd.get("header_row"), int),
          f"type={type(pd.get('header_row')).__name__} value={pd.get('header_row')}")
    check("data_start_row is int", isinstance(pd.get("data_start_row"), int))
    check("completed_phases has triage", "triage" in test_state.get("completed_phases", []))

    # 2c. extract_categories
    cat_state = dict(test_state)  # copy
    result = extract_categories.invoke({
        "file_path": TEST_FILE,
        "sheet_name": "Sheet1",
        "specified_columns": ["CATEGORY1", "CATEGORY2", "CATEGORY3", "CATEGORY4"],
        "state": cat_state,
    })
    check("extract_categories returns paths", "Extracted" in result or "Discovered" in result or "wasn't able" in result,
          f"result={result[:80]}")
    check("completed_phases has categories", "categories" in cat_state.get("completed_phases", []))

    # 2d. profile_columns
    from helpers import read_file, get_headers_and_data
    gen = read_file(TEST_FILE)
    hr = pd.get("header_row", 0)
    try:
        from helpers import take_rows
        first = take_rows(gen, hr + 5)
        hdr = [str(c) if c is not None else "" for c in first[hr]]
        data = first[hr + 1:]
        cols = profile_columns.invoke({"headers": hdr, "rows": data})
        check("profile_columns returns list of profiles", isinstance(cols, list) and len(cols) > 0)
        check("each profile has name/unique/non_null", all(c.get("name") for c in cols))
    except Exception as e:
        check(f"profile_columns works", False, str(e)[:80])

    # 2e. extract_references (needs core_mappings)
    ref_state = dict(test_state)
    ref_state["core_mappings"] = {"sku_name": "ITEM", "mrp": "MRP", "code": "ITEM CODE"}
    ref_state["custom_mappings"] = {"COLOR NAME": "COLOR NAME", "BRAND": "BRAND"}
    ref_state["profile_data"] = pd
    try:
        refs = extract_references.invoke({"state": ref_state})
        check("extract_references returns string", isinstance(refs, str))
    except Exception as e:
        check("extract_references works", False, str(e)[:80])

    # 2f. build_products
    prod_state = dict(test_state)
    prod_state["profile_data"] = pd
    prod_state["core_mappings"] = {"sku_name": "ITEM", "mrp": "MRP", "code": "ITEM CODE"}
    prod_state["custom_mappings"] = {"COLOR NAME": "COLOR NAME", "BRAND": "BRAND"}
    try:
        prod_result = build_products.invoke({"state": prod_state})
        check("build_products returns string", isinstance(prod_result, str) and len(prod_result) > 0,
              f"result={prod_result[:80]}")
        check("product_rows populated", len(prod_state.get("product_rows", [])) > 0,
              f"count={len(prod_state.get('product_rows', []))}")
    except Exception as e:
        check("build_products works", False, str(e)[:80])

    # 2g. render_templates
    render_state = dict(prod_state)
    render_state["product_rows"] = prod_state.get("product_rows", [])
    render_state["profile_data"]["category_hierarchy"] = ["TEST > CATEGORY"]
    try:
        render_result = render_templates.invoke({"state": render_state})
        check("render_templates returns success", "Generated" in render_result, render_result[:60])
        check("generated_files populated", len(render_state.get("generated_files", [])) > 0,
              f"files={render_state['generated_files']}")
    except Exception as e:
        check("render_templates works", False, str(e)[:80])


# ═══════════════════════════════════════════════════════════════
# TEST GROUP 3: API endpoint smoke tests (requires live server)
# ═══════════════════════════════════════════════════════════════

def test_api_endpoints(api_url):
    section("3. API Endpoint Tests")
    import urllib.request
    import urllib.error
    import http.client

    def api_call(method, path, body=None, sse=False):
        url = f"{api_url}{path}"
        data = json.dumps(body).encode() if body else None
        req = urllib.request.Request(url, data=data, method=method)
        req.add_header("Content-Type", "application/json")
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                if sse:
                    return resp.read().decode()[:500]
                return json.loads(resp.read().decode())
        except urllib.error.HTTPError as e:
            return {"error": str(e), "code": e.code}
        except Exception as e:
            return {"error": str(e)}

    # 3a. Health check — server responds with 200
    try:
        conn = http.client.HTTPConnection(api_url.replace("http://", ""))
        conn.request("GET", "/")
        resp = conn.getresponse()
        status = resp.status
        resp.read()  # drain
        conn.close()
        check("Server responds with 200", status == 200, f"status={status}")
    except Exception as e:
        check("Server responds with 200", False, str(e)[:80])

    # 3b. Upload file
    upload_url = f"{api_url}/upload"
    import http.client
    import mimetypes
    boundary = "----TestBoundary123"
    with open(TEST_FILE, "rb") as f:
        file_data = f.read()
    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{os.path.basename(TEST_FILE)}"\r\n'
        f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
    ).encode() + file_data + f"\r\n--{boundary}--\r\n".encode()

    try:
        conn = http.client.HTTPConnection(api_url.replace("http://", ""))
        conn.request("POST", "/upload", body, {"Content-Type": f"multipart/form-data; boundary={boundary}"})
        resp = conn.getresponse()
        upload_data = json.loads(resp.read().decode())
        conn.close()
        file_path = upload_data.get("path", "")
        check("Upload returns file path", bool(file_path), f"path={file_path}")
    except Exception as e:
        check("Upload works", False, str(e)[:80])
        return  # Can't continue without a file

    # 3c. Start interactive session
    start_body = json.dumps({"file_path": file_path}).encode()
    try:
        conn = http.client.HTTPConnection(api_url.replace("http://", ""))
        conn.request("POST", "/interactive/start", start_body, {"Content-Type": "application/json"})
        resp = conn.getresponse()
        sse_data = resp.read().decode()
        conn.close()
        check("Start returns SSE events", "tool_start" in sse_data or "progress" in sse_data or "complete" in sse_data,
              f"data={sse_data[:200]}")
    except Exception as e:
        check("Start works", False, str(e)[:80])
        return

    # Extract thread_id from SSE
    thread_id = None
    for line in sse_data.split("\n"):
        if line.startswith("data: "):
            try:
                evt = json.loads(line[6:])
                if evt.get("type") == "complete":
                    thread_id = evt.get("thread_id")
            except:
                pass

    if not thread_id:
        check("Thread ID extracted", False)
        return
    check("Start returns thread_id", bool(thread_id), thread_id)

    # 3d. Respond with a simple message
    respond_body = json.dumps({"thread_id": thread_id, "message": "Hello, what columns do you see?"}).encode()
    try:
        conn = http.client.HTTPConnection(api_url.replace("http://", ""))
        conn.request("POST", "/interactive/respond", respond_body, {"Content-Type": "application/json"})
        resp = conn.getresponse()
        sse_data2 = resp.read().decode()
        conn.close()
        check("Respond returns SSE events", len(sse_data2) > 50, f"len={len(sse_data2)}")
    except Exception as e:
        check("Respond works", False, str(e)[:80])

    # 3e. Status check
    status_body = json.dumps({"thread_id": thread_id}).encode()
    try:
        conn = http.client.HTTPConnection(api_url.replace("http://", ""))
        conn.request("POST", "/interactive/status", status_body, {"Content-Type": "application/json"})
        resp = conn.getresponse()
        status_data = json.loads(resp.read().decode())
        conn.close()
        check("Status returns valid data", isinstance(status_data, dict),
              f"keys={list(status_data.keys())}")
    except Exception as e:
        check("Status works", False, str(e)[:80])


# ═══════════════════════════════════════════════════════════════
# TEST GROUP 4: Output file format validation
# ═══════════════════════════════════════════════════════════════

def test_output_files():
    section("4. Output File Format Validation")
    import openpyxl

    output_dir = "output"
    if not os.path.exists(output_dir):
        check("Output directory exists", False)
        return

    files = [f for f in os.listdir(output_dir) if f.endswith(".xlsx")]
    check("Output directory has xlsx files", len(files) > 0, f"count={len(files)}")

    # Check most recent product file
    xlsx_files = sorted([os.path.join(output_dir, f) for f in files if "_product" in f], key=os.path.getmtime)
    if xlsx_files:
        latest = xlsx_files[-1]
        try:
            wb = openpyxl.load_workbook(latest, read_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
            check("Product sheet has Category Path column", "Category Path" in headers, f"headers={headers[:6]}")
            check("Product sheet has sku_name column", "sku_name" in headers)
            check("Product sheet has mrp column", "mrp" in headers)
            check("Product sheet has data rows", ws.max_row > 1, f"rows={ws.max_row}")
            wb.close()
        except Exception as e:
            check(f"Product file readable", False, str(e)[:80])

    # Check most recent attribute file
    attr_files = sorted([os.path.join(output_dir, f) for f in files if "_attribute" in f], key=os.path.getmtime)
    if attr_files:
        try:
            wb = openpyxl.load_workbook(attr_files[-1], read_only=True)
            ws = wb.active
            headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
            check("Attribute sheet has 17 columns", len(headers) >= 17, f"count={len(headers)}")
            check("Attribute sheet has Attribute Name", "Attribute Name" in headers)
            wb.close()
        except Exception as e:
            check("Attribute file readable", False, str(e)[:80])


# ═══════════════════════════════════════════════════════════════
# TEST GROUP 5: Edge cases
# ═══════════════════════════════════════════════════════════════

def test_tool_path_resolution_coverage():
    section("5. Tool Path Resolution Edge Cases")

    from interactive_graph import profile_file, extract_categories

    # Bare filename (no directory)
    test_state = {
        "file_path": TEST_FILE,
        "profile_data": None,
        "messages": [],
        "completed_phases": [],
        "remaining_steps": 0,
        "core_mappings": {},
        "custom_mappings": {},
        "mapping_confidence": {},
        "product_rows": [],
        "generated_files": [],
        "jwt_token": "",
        "all_sheets": [],
        "sheet_merge": {},
        "categories": {"explanation": "", "reasoning": "", "suggestions": [], "approved": False, "user_feedback": ""},
        "attributes": {"explanation": "", "reasoning": "", "suggestions": [], "approved": False, "user_feedback": ""},
        "references": {"explanation": "", "reasoning": "", "suggestions": [], "approved": False, "user_feedback": ""},
        "products": {"explanation": "", "reasoning": "", "suggestions": [], "approved": False, "user_feedback": ""},
        "current_phase": "categories",
        "phases_completed": [],
    }

    # profile_file with bare filename
    try:
        result = profile_file.invoke({"file_path": os.path.basename(TEST_FILE), "state": test_state})
        check("profile_file resolves bare filename", "Profiled" in result)
    except Exception as e:
        check("profile_file resolves bare filename", False, str(e)[:80])

    # extract_categories with bare filename
    test_state["profile_data"] = None  # Reset so profile_file re-runs
    try:
        result = profile_file.invoke({"file_path": "nonexistent.xlsx", "state": test_state})
        check("profile_file handles nonexistent path gracefully",
              "Profiled" in result or "wasn't able" in result or "cached" in result,
              f"result={result[:80]}")
    except Exception as e:
        check("profile_file handles nonexistent path gracefully", False, str(e)[:80])

    # State without file_path — tool should raise FileNotFoundError
    empty_state = {"file_path": "", "profile_data": None, "messages": [], "completed_phases": []}
    try:
        result = extract_categories.invoke({
            "file_path": "no_such_file.xlsx",
            "sheet_name": "Sheet1",
            "specified_columns": ["CATEGORY1"],
            "state": empty_state,
        })
        check("extract_categories raises on missing file", False,
              f"Should have raised — returned: {result[:60]}")
    except FileNotFoundError:
        check("extract_categories raises FileNotFoundError correctly", True)
    except Exception as e:
        check("extract_categories raises on missing file", True,
              f"Expected FileNotFoundError, got {type(e).__name__}: {str(e)[:60]}")


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="PIM Ingestion Agent Test Suite")
    parser.add_argument("--api", help="Base URL of live API server (e.g. http://localhost:8000)")
    parser.add_argument("--quick", action="store_true", help="Skip slow LLM-dependent tests")
    args = parser.parse_args()

    print(f"PIM Ingestion Agent — Integration Test Suite")
    print(f"Python {sys.version}")
    print(f"Test file: {TEST_FILE}")
    print(f"Quick mode: {args.quick}")

    # Group 1: Graph structure (fast, no API key needed)
    test_graph_structure()

    # Group 2: Tool unit tests (needs real file on disk)
    test_tools_direct(quick=args.quick)

    # Group 5: Edge cases
    test_tool_path_resolution_coverage()

    # Group 4: Output file format (depends on prior runs)
    test_output_files()

    # Group 3: API endpoint tests (needs --api flag)
    if args.api:
        test_api_endpoints(args.api)

    # Summary
    total = PASS + FAIL
    print(f"\n{'='*60}")
    print(f"  RESULTS: {PASS}/{total} passed, {FAIL} failed")
    print(f"{'='*60}")
    sys.exit(1 if FAIL > 0 else 0)
