import os
import math
import logging
import warnings
import polars as pl

logger = logging.getLogger("pim_data_plane")
logging.getLogger("polars").setLevel(logging.ERROR)
warnings.filterwarnings("ignore", message="Could not determine dtype")


def _detect_sheet(path, preferred=None):
    if preferred:
        return preferred
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        best = max(wb.sheetnames, key=lambda sn: (
            wb[sn].max_row or 0) * (wb[sn].max_column or 0))
        wb.close()
        return best
    except Exception:
        return 0


def _lazy_or_read(path):
    """Return a LazyFrame regardless of file format (CSV or xlsx)."""
    if path.endswith(".csv"):
        return pl.scan_csv(path)
    df = pl.read_excel(path, engine="calamine", has_header=True)
    return df.lazy()


def get_variance_sample_polars(source_path: str, title_col: str, limit: int = 5) -> list[dict]:
    """
    Stage 1 & 2: Identifies a candidate block of similar products without OOM.
    Groups by the first 2 words of the product title to simulate LSH blocking.
    """
    lf = _lazy_or_read(source_path)

    if title_col not in lf.collect_schema().names():
        return []

    block_expr = (
        pl.col(title_col)
        .cast(pl.Utf8)
        .str.split(" ")
        .list.slice(0, 2)
        .list.join(" ")
        .alias("block_key")
    )

    sample_df = (
        lf.with_columns(block_expr)
        .filter(pl.col("block_key").is_not_null())
        .filter(pl.len().over("block_key") > 1)
        .head(limit)
        .collect(streaming=True)
    )

    return sample_df.to_dicts()


def execute_variant_rules_polars(source_path: str, invariants: list[str], variants: list[str], output_path: str):
    """
    Executes the LLM-calibrated invariant/variant rules over 1,000,000 rows.
    """
    lf = _lazy_or_read(source_path)

    valid_invariants = [col for col in invariants if col in lf.collect_schema().names()]

    if valid_invariants:
        parent_expr = pl.concat_str([pl.col(k).cast(pl.Utf8) for k in valid_invariants]).hash().cast(pl.Utf8)
    else:
        parent_expr = pl.lit("")

    variant_attr_str = "::".join(variants)

    lf_enriched = lf.with_columns([
        parent_expr.alias("parent_sku"),
        pl.lit(variant_attr_str).alias("variant_attributes")
    ])

    lf_enriched.sink_csv(output_path)


def detect_header_row_and_headers(path, sheet_name=None):
    ext = os.path.splitext(path)[1].lower()
    raw_rows = []

    if ext == ".csv":
        import charset_normalizer
        result = charset_normalizer.from_path(path).best()
        enc = result.encoding if result else "utf-8"
        with open(path, "r", encoding=enc) as f:
            for _ in range(20):
                line = f.readline()
                if not line:
                    break
                raw_rows.append(line.rstrip("\n").split(","))
    else:
        import openpyxl
        sn = sheet_name if sheet_name else _detect_sheet(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sn]
        raw_rows = [list(r) for _, r in zip(range(20), ws.iter_rows(values_only=True))]
        wb.close()

    from agents import detect_header_via_llm
    try:
        header_row, data_start_row = detect_header_via_llm(raw_rows)
    except Exception as e:
        logger.warning(f"header detection failed: {e}")
        header_row, data_start_row = 0, 1

    if not raw_rows or header_row >= len(raw_rows):
        return 0, 1, []

    if data_start_row < header_row + 1:
        data_start_row = header_row + 1

    headers = [str(c) if c is not None else ""
               for c in raw_rows[header_row]]

    return header_row, data_start_row, headers


def get_lazy_frame(path, sheet_name=None, header_row=0, headers=None):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")

    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        import charset_normalizer
        result = charset_normalizer.from_path(path).best()
        detected_enc = result.encoding if result else "utf-8"
        if detected_enc.lower() not in ("utf8", "utf-8", "utf-8-sig"):
            import shutil
            tmp = path + ".utf8.tmp"
            with open(path, "r", encoding=detected_enc) as src:
                with open(tmp, "w", encoding="utf-8") as dst:
                    shutil.copyfileobj(src, dst)
            path = tmp
        return pl.scan_csv(path, encoding="utf8", skip_rows=header_row + 1,
                           has_header=False, new_columns=headers,
                           infer_schema_length=1000, ignore_errors=True,
                           null_values=["", "NA", "N/A", "NULL", "null"])

    if ext == ".xls":
        import xlrd
        xl = xlrd.open_workbook(path)
        sn = sheet_name if sheet_name else xl.sheet_names()[0]
        ws = xl.sheet_by_name(sn)
        final_headers = headers or []
        data = []
        for r in range(header_row + 1, ws.nrows):
            data.append([str(ws.cell(r, c).value).strip()
                        for c in range(ws.ncols)])
        df = pl.DataFrame(data, schema=final_headers or None, orient="row")
        return df.lazy()

    sn = sheet_name if sheet_name else _detect_sheet(path)
    df = pl.read_excel(path, sheet_name=sn, engine="calamine",
                       has_header=False,
                       read_options={"skip_rows": header_row + 1})
    if headers:
        clean = []
        seen = {}
        for h in headers[:len(df.columns)]:
            if h in seen:
                seen[h] += 1
                clean.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                clean.append(h)
        df.columns = clean
    return df.lazy()


def profile_large_file(path, sheet_name=None):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")

    header_row, data_start_row, headers = detect_header_row_and_headers(
        path, sheet_name)

    if not headers:
        return {"headers": [], "profiles": [], "row_count": 0, "sample_rows": []}

    lazy = get_lazy_frame(path, sheet_name, header_row, headers)
    schema = lazy.collect_schema()
    col_names = schema.names()

    stats = lazy.select([
        pl.all().null_count().name.suffix("_nulls"),
        pl.all().n_unique().name.suffix("_uniques"),
    ]).collect(streaming=True)

    profiles = []
    for col in col_names:
        null_count = stats[f"{col}_nulls"][0]
        unique_count = stats[f"{col}_uniques"][0]
        sample_values = lazy.select(
            pl.col(col).drop_nulls().limit(5)
        ).collect()[col].to_list()
        sample_values_clean = [str(v) if v is not None else "" for v in sample_values]
        profiles.append({
            "column_name": col,
            "null_count": int(null_count),
            "unique_count": int(unique_count),
            "sample_values": sample_values_clean,
            "data_type": str(schema[col]),
        })

    row_count = lazy.select(pl.len()).collect()[0, 0]
    sample_rows_df = lazy.limit(3).collect()
    sample_rows = sample_rows_df.to_dicts()

    return {
        "headers": headers,
        "profiles": profiles,
        "row_count": int(row_count),
        "sample_rows": sample_rows,
    }


def extract_categories_polars(path, columns, sheet_name=None):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")

    header_row, _, headers = detect_header_row_and_headers(path, sheet_name)
    cols_to_use = [c.strip() for c in columns if c.strip() in headers]

    if not cols_to_use:
        return []

    lazy = get_lazy_frame(path, sheet_name, header_row, headers)
    df_categories = lazy.select(cols_to_use).unique().collect(streaming=True)
    paths = set()

    for row in df_categories.iter_rows():
        nodes = []
        for val in row:
            if val is None:
                continue
            if isinstance(val, float) and math.isnan(val):
                continue
            s = str(val).strip()
            if not s or s.lower() == "nan":
                continue
            nodes.append(s)
        if nodes:
            paths.add(" > ".join(nodes))

    return sorted(list(paths))


def stream_product_export_polars(source_path, core_mappings, output_path,
                                 sheet_name=None):
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"File not found: {source_path}")

    header_row, _, headers = detect_header_row_and_headers(source_path, sheet_name)
    lazy = get_lazy_frame(source_path, sheet_name, header_row, headers)

    projection = {}
    for target_pim, source_col in core_mappings.items():
        if source_col in headers:
            projection[target_pim] = pl.col(source_col)

    if not projection:
        raise ValueError(
            f"None of {len(core_mappings)} mapped columns exist in source headers")

    lazy_mapped = lazy.select(**projection)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    if output_path.endswith(".csv"):
        lazy_mapped.sink_csv(output_path)
    else:
        df_mapped = lazy_mapped.collect(streaming=True)
        df_mapped.write_excel(output_path)
